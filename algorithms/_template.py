# Referans şablon — AI ürettiği her algoritma bu imzaya uymalıdır.
#
# Bu dosyayı silmeyin. Yeni algoritma dosyaları func_{id}.py olarak
# bu dizinde otomatik oluşturulur.

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import math


def create_excel(data: dict, output_path: str) -> None:
    """
    Şablon create_excel fonksiyonu.

    Args:
        data: {"satirlar": [{"Sütun1": değer, "Sütun2": değer, ...}, ...]}
        output_path: Çıktı dosya yolu (.xlsx)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Veri"

    satirlar = data.get("satirlar", [])
    if not satirlar:
        wb.save(output_path)
        return

    # Başlıklar
    headers = list(satirlar[0].keys())
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Veri satırları
    for row_idx, satir in enumerate(satirlar, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=satir.get(header))

    # Otomatik genişlik
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(satirlar) + 2)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A2"
    wb.save(output_path)
