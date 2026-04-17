import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os.path
import math

def create_excel(data: dict, output_path: str) -> None:
    """
    Turizm ajansı oda yerleşim planını profesyonel Excel formatında oluşturur.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Oda_Yerlesim_Plani"

    # Sütun Tanımları
    columns = [
        "GRUP ADI / ODA NO",
        "SOYİSİM / ODA TİPİ",
        "TC-PASAPORT / MANZARA",
        "CİNSİYET / ÖZEL NOT",
        "YAŞ",
        "GRUP ID"
    ]

    # Stil Tanımları
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    summary_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    summary_font = Font(bold=True)
    
    border_medium = Border(
        left=Side(style='medium', color='999999'),
        right=Side(style='medium', color='999999'),
        top=Side(style='medium', color='999999'),
        bottom=Side(style='medium', color='999999')
    )
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Başlık Satırını Yaz
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_medium

    # Veri Satırlarını Yaz
    rows = data.get("satirlar", [])
    current_row = 2
    
    # Sayaçlar (Özet için)
    total_people = 0
    rooms_set = set()
    waitlist_count = 0

    for idx, item in enumerate(rows):
        is_even = (current_row % 2 == 0)
        row_fill = even_row_fill if is_even else odd_row_fill
        
        # Veriyi işle ve hücrelere yaz
        col_values = [
            item.get("GRUP ADI / ODA NO"),
            item.get("SOYİSİM / ODA TİPİ"),
            item.get("TC-PASAPORT / MANZARA"),
            item.get("CİNSİYET / ÖZEL NOT"),
            item.get("YAŞ"),
            item.get("GRUP ID")
        ]

        # İstatistik Takibi
        val_first = str(col_values[0] or "")
        val_age = col_values[4]
        
        if "ODA" in val_first.upper() and "UYARI" not in val_first.upper():
            rooms_set.add(val_first)
        if "UYARI" in val_first.upper() or "bekleme" in str(item.get("CİNSİYET / ÖZEL NOT", "")).lower():
            waitlist_count += 1
        if val_age is not None and str(val_age).isdigit():
            total_people += 1

        for col_idx, value in enumerate(col_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = border_thin
            
            # Hizalama ve Formatlama
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            
            # Özel durum: Uyarı satırları kırmızı font
            if "UYARI" in str(value).upper():
                cell.font = Font(color="FF0000", bold=True)

        current_row += 1

    # Özet Satırları
    current_row += 1
    summary_labels = [
        ("Toplam Kişi Sayısı", total_people),
        ("Dolu Oda Sayısı", len(rooms_set)),
        ("Bekleme Listesindeki Kişi Sayısı", waitlist_count)
    ]

    for label, val in summary_labels:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        label_cell = ws.cell(row=current_row, column=1, value=label)
        val_cell = ws.cell(row=current_row, column=3, value=val)
        
        label_cell.font = summary_font
        label_cell.fill = summary_fill
        val_cell.font = summary_font
        val_cell.fill = summary_fill
        val_cell.alignment = Alignment(horizontal="right")
        
        current_row += 1

    # Sütun Genişliklerini Ayarla
    for col_idx in range(1, len(columns) + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in ws[column]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)

    # Sayfa Ayarları
    ws.freeze_panes = "A2"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1

    wb.save(output_path)