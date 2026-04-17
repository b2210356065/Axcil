from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os
import re


def _get_default_data():
    """
    Grand Paradise Resort - 15-22 Temmuz 2026 rezervasyon verisi.
    Oda yerleştirme kuralları uygulanmış, Excel'e yazılmaya hazır satirlar.

    Odalar:
      2-kisilik: 101-200 (100 oda)
      3-kisilik: 201-280 (80 oda)
      4-kisilik: 281-330 (50 oda)
    """
    satirlar = []

    def oda(no, tip, manzara=None, ozel_not=None):
        return {
            "Satır Tipi": "ODA_BAŞLIK",
            "Oda No": no, "Oda Tipi": tip, "Manzara": manzara or "Standart",
            "Özel Not (Oda)": ozel_not,
            "İsim": None, "Soyisim": None, "TC/Pasaport": None,
            "Cinsiyet": None, "Yaş": None, "Grup ID": None,
            "İlişki Tipi": None, "Oda Tercihi": None,
            "Durum": None, "Uyarı Açıklama": None,
        }

    def misafir(isim, soyisim, tc, cinsiyet, yas, grup, iliski, tercih,
                durum="Yerleştirildi", uyari=None):
        return {
            "Satır Tipi": "MİSAFİR",
            "Oda No": None, "Oda Tipi": None, "Manzara": None, "Özel Not (Oda)": None,
            "İsim": isim, "Soyisim": soyisim, "TC/Pasaport": tc,
            "Cinsiyet": cinsiyet, "Yaş": yas, "Grup ID": grup,
            "İlişki Tipi": iliski, "Oda Tercihi": tercih,
            "Durum": durum, "Uyarı Açıklama": uyari,
        }

    def uyari_satir(isim, soyisim, tc, cinsiyet, yas, grup, iliski, tercih, aciklama):
        return {
            "Satır Tipi": "UYARI",
            "Oda No": None, "Oda Tipi": None, "Manzara": None, "Özel Not (Oda)": None,
            "İsim": isim, "Soyisim": soyisim, "TC/Pasaport": tc,
            "Cinsiyet": cinsiyet, "Yaş": yas, "Grup ID": grup,
            "İlişki Tipi": iliski, "Oda Tercihi": tercih,
            "Durum": "Bekleme Listesi", "Uyarı Açıklama": aciklama,
        }

    # ----------------------------------------------------------------
    # ODA 281 — 4 Kişilik | GRUP-1 Yılmaz Ailesi
    # ----------------------------------------------------------------
    satirlar.append(oda(281, "4 Kişilik"))
    satirlar.append(misafir("Mehmet",  "Yılmaz", "12345678901", "E", 45, "GRUP-1", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Ayşe",    "Yılmaz", "12345678902", "K", 42, "GRUP-1", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Can",     "Yılmaz", "12345678903", "E", 18, "GRUP-1", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Elif",    "Yılmaz", "12345678904", "K", 15, "GRUP-1", "Aile", "4 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 282 — 4 Kişilik | GRUP-2 Demir Ailesi (5 kişi → 4+1)
    # ----------------------------------------------------------------
    satirlar.append(oda(282, "4 Kişilik"))
    satirlar.append(misafir("Ahmet",  "Demir", "23456789012", "E", 50, "GRUP-2", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Fatma",  "Demir", "23456789013", "K", 48, "GRUP-2", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Ali",    "Demir", "23456789014", "E", 22, "GRUP-2", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Zeynep", "Demir", "23456789015", "K", 19, "GRUP-2", "Aile", "4 Kişilik"))

    # ODA 101 — 2 Kişilik | GRUP-2 Demir (kalan 1 kişi)
    satirlar.append(oda(101, "2 Kişilik"))
    satirlar.append(misafir("Ece", "Demir", "23456789016", "K", 16, "GRUP-2", "Aile", "4 Kişilik",
                            uyari="Gruptan ayrı odada (5 kişi → 4+1 bölünme)"))

    # ----------------------------------------------------------------
    # ODA 201 — 3 Kişilik | GRUP-3 Kaya Ailesi
    # ----------------------------------------------------------------
    satirlar.append(oda(201, "3 Kişilik"))
    satirlar.append(misafir("Murat", "Kaya", "34567890123", "E", 38, "GRUP-3", "Aile", "3 Kişilik"))
    satirlar.append(misafir("Selin", "Kaya", "34567890124", "K", 35, "GRUP-3", "Aile", "3 Kişilik"))
    satirlar.append(misafir("Deniz", "Kaya", "34567890125", "K", 12, "GRUP-3", "Aile", "3 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 102 — 2 Kişilik | GRUP-4 Çelik Ailesi (Evli çift)
    # ----------------------------------------------------------------
    satirlar.append(oda(102, "2 Kişilik"))
    satirlar.append(misafir("Emre", "Çelik", "45678901234", "E", 30, "GRUP-4", "Evli", "2 Kişilik"))
    satirlar.append(misafir("Buse", "Çelik", "45678901235", "K", 28, "GRUP-4", "Evli", "2 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 283 — 4 Kişilik | GRUP-5 Arslan Ailesi
    # ----------------------------------------------------------------
    satirlar.append(oda(283, "4 Kişilik"))
    satirlar.append(misafir("Hakan", "Arslan", "56789012345", "E", 43, "GRUP-5", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Gül",   "Arslan", "56789012346", "K", 41, "GRUP-5", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Burak", "Arslan", "56789012347", "E", 17, "GRUP-5", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Merve", "Arslan", "56789012348", "K", 14, "GRUP-5", "Aile", "4 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 284 — 4 Kişilik | GRUP-6 Erkek Arkadaş
    # ----------------------------------------------------------------
    satirlar.append(oda(284, "4 Kişilik"))
    satirlar.append(misafir("Cem",    "Öztürk", "67890123456", "E", 25, "GRUP-6", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Barış",  "Koç",    "67890123457", "E", 26, "GRUP-6", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Tolga",  "Şahin",  "67890123458", "E", 25, "GRUP-6", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Serkan", "Polat",  "67890123459", "E", 27, "GRUP-6", "Arkadaş", "4 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 202 — 3 Kişilik | GRUP-7 Erkek Arkadaş
    # ----------------------------------------------------------------
    satirlar.append(oda(202, "3 Kişilik"))
    satirlar.append(misafir("Oğuz", "Yıldız", "78901234567", "E", 24, "GRUP-7", "Arkadaş", "3 Kişilik"))
    satirlar.append(misafir("Kaan", "Aydın",  "78901234568", "E", 24, "GRUP-7", "Arkadaş", "3 Kişilik"))
    satirlar.append(misafir("Eren", "Bulut",  "78901234569", "E", 23, "GRUP-7", "Arkadaş", "3 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 103 — 2 Kişilik | GRUP-8 Erkek Arkadaş
    # ----------------------------------------------------------------
    satirlar.append(oda(103, "2 Kişilik"))
    satirlar.append(misafir("Volkan", "Kurt",  "89012345678", "E", 29, "GRUP-8", "Arkadaş", "2 Kişilik"))
    satirlar.append(misafir("Onur",   "Özkan", "89012345679", "E", 28, "GRUP-8", "Arkadaş", "2 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 285 — 4 Kişilik | GRUP-9 Kadın Arkadaş
    # ----------------------------------------------------------------
    satirlar.append(oda(285, "4 Kişilik"))
    satirlar.append(misafir("Seda",   "Türk",  "90123456780", "K", 26, "GRUP-9", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Canan",  "Aksoy", "90123456781", "K", 27, "GRUP-9", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Pınar",  "Güler", "90123456782", "K", 25, "GRUP-9", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Damla",  "Taş",   "90123456783", "K", 26, "GRUP-9", "Arkadaş", "4 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 203 — 3 Kişilik | GRUP-10 Kadın Arkadaş
    # ----------------------------------------------------------------
    satirlar.append(oda(203, "3 Kişilik"))
    satirlar.append(misafir("İrem",  "Kaplan", "01234567891", "K", 23, "GRUP-10", "Arkadaş", "3 Kişilik"))
    satirlar.append(misafir("Begüm", "Doğan",  "01234567892", "K", 24, "GRUP-10", "Arkadaş", "3 Kişilik"))
    satirlar.append(misafir("Naz",   "Şen",    "01234567893", "K", 23, "GRUP-10", "Arkadaş", "3 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 104 — 2 Kişilik | GRUP-11 Kadın Arkadaş
    # ----------------------------------------------------------------
    satirlar.append(oda(104, "2 Kişilik"))
    satirlar.append(misafir("Ebru",  "Erdem", "11234567894", "K", 30, "GRUP-11", "Arkadaş", "2 Kişilik"))
    satirlar.append(misafir("Gizem", "Aslan", "11234567895", "K", 29, "GRUP-11", "Arkadaş", "2 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 286 — 4 Kişilik | GRUP-12 İki Evli Çift
    # ----------------------------------------------------------------
    satirlar.append(oda(286, "4 Kişilik"))
    satirlar.append(misafir("Kerem", "Öz",    "21234567896", "E", 32, "GRUP-12", "Evli", "4 Kişilik"))
    satirlar.append(misafir("Derya", "Öz",    "21234567897", "K", 30, "GRUP-12", "Evli", "4 Kişilik"))
    satirlar.append(misafir("Sinan", "Tekin", "21234567898", "E", 33, "GRUP-12", "Evli", "4 Kişilik"))
    satirlar.append(misafir("Burcu", "Tekin", "21234567899", "K", 31, "GRUP-12", "Evli", "4 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 287 — 4 Kişilik | GRUP-13 Üç Çift (4+2 bölünme)
    # ----------------------------------------------------------------
    satirlar.append(oda(287, "4 Kişilik"))
    satirlar.append(misafir("Deniz",  "Sarı",  "31234567800", "E", 28, "GRUP-13", "Evli", "4 Kişilik"))
    satirlar.append(misafir("Esra",   "Sarı",  "31234567801", "K", 27, "GRUP-13", "Evli", "4 Kişilik"))
    satirlar.append(misafir("Koray",  "Beyaz", "31234567802", "E", 29, "GRUP-13", "Evli", "4 Kişilik"))
    satirlar.append(misafir("Duygu",  "Beyaz", "31234567803", "K", 28, "GRUP-13", "Evli", "4 Kişilik"))

    # ODA 105 — 2 Kişilik | GRUP-13 (kalan çift)
    satirlar.append(oda(105, "2 Kişilik"))
    satirlar.append(misafir("Umut",  "Kılıç", "31234567804", "E", 30, "GRUP-13", "Evli", "4 Kişilik",
                            uyari="Gruptan ayrı odada (6 kişi → 4+2 bölünme)"))
    satirlar.append(misafir("Aylin", "Kılıç", "31234567805", "K", 29, "GRUP-13", "Evli", "4 Kişilik",
                            uyari="Gruptan ayrı odada (6 kişi → 4+2 bölünme)"))

    # ----------------------------------------------------------------
    # ODA 204 — 3 Kişilik | Tek Rezervasyon Erkekler
    # ----------------------------------------------------------------
    satirlar.append(oda(204, "3 Kişilik"))
    satirlar.append(misafir("Mustafa", "Uzun",  "41234567806", "E", 35, "TEKK-1", "Yalnız", "2 Kişilik",
                            uyari="Tek rezervasyon — aynı cinsiyetten bireylerle eşleştirildi"))
    satirlar.append(misafir("Hüseyin", "Güneş", "51234567807", "E", 40, "TEKK-2", "Yalnız", "2 Kişilik",
                            uyari="Tek rezervasyon — aynı cinsiyetten bireylerle eşleştirildi"))
    satirlar.append(misafir("Ramazan", "Çınar", "61234567808", "E", 38, "TEKK-3", "Yalnız", "2 Kişilik",
                            uyari="Tek rezervasyon — aynı cinsiyetten bireylerle eşleştirildi"))

    # ----------------------------------------------------------------
    # ODA 205 — 3 Kişilik | Tek Rezervasyon Kadınlar
    # ----------------------------------------------------------------
    satirlar.append(oda(205, "3 Kişilik"))
    satirlar.append(misafir("Aysel",  "Bozkurt", "71234567809", "K", 33, "TEKK-4", "Yalnız", "2 Kişilik",
                            uyari="Tek rezervasyon — aynı cinsiyetten bireylerle eşleştirildi"))
    satirlar.append(misafir("Hacer",  "Yavuz",   "81234567810", "K", 36, "TEKK-5", "Yalnız", "2 Kişilik",
                            uyari="Tek rezervasyon — aynı cinsiyetten bireylerle eşleştirildi"))
    satirlar.append(misafir("Nurcan", "Ak",       "91234567811", "K", 34, "TEKK-6", "Yalnız", "2 Kişilik",
                            uyari="Tek rezervasyon — aynı cinsiyetten bireylerle eşleştirildi"))

    # ----------------------------------------------------------------
    # ODA 288 — 4 Kişilik | GRUP-14 Büyük Aile (7 kişi → 4+3)
    # ----------------------------------------------------------------
    satirlar.append(oda(288, "4 Kişilik"))
    satirlar.append(misafir("İbrahim", "Şimşek", "10234567812", "E", 65, "GRUP-14", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Hatice",  "Şimşek", "10234567813", "K", 62, "GRUP-14", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Recep",   "Şimşek", "10234567814", "E", 40, "GRUP-14", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Emine",   "Şimşek", "10234567815", "K", 38, "GRUP-14", "Aile", "4 Kişilik"))

    # ODA 206 — 3 Kişilik | GRUP-14 (kalan 3 kişi)
    satirlar.append(oda(206, "3 Kişilik"))
    satirlar.append(misafir("Mehmet Ali", "Şimşek", "10234567816", "E", 20, "GRUP-14", "Aile", "4 Kişilik",
                            uyari="Gruptan ayrı odada (7 kişi → 4+3 bölünme)"))
    satirlar.append(misafir("Fatma Nur",  "Şimşek", "10234567817", "K", 18, "GRUP-14", "Aile", "4 Kişilik",
                            uyari="Gruptan ayrı odada (7 kişi → 4+3 bölünme)"))
    satirlar.append(misafir("Ahmet Can",  "Şimşek", "10234567818", "E", 15, "GRUP-14", "Aile", "4 Kişilik",
                            uyari="Gruptan ayrı odada (7 kişi → 4+3 bölünme)"))

    # ----------------------------------------------------------------
    # ODA 289 — 4 Kişilik | GRUP-15 Acar Ailesi
    # ----------------------------------------------------------------
    satirlar.append(oda(289, "4 Kişilik"))
    satirlar.append(misafir("Selim",  "Acar", "20234567819", "E", 45, "GRUP-15", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Nazan",  "Acar", "20234567820", "K", 43, "GRUP-15", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Berkay", "Acar", "20234567821", "E", 19, "GRUP-15", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Yağmur", "Acar", "20234567822", "K", 16, "GRUP-15", "Aile", "4 Kişilik"))

    # ODA 290 — 4 Kişilik | GRUP-15 Vural Ailesi (akraba, yan yana)
    satirlar.append(oda(290, "4 Kişilik"))
    satirlar.append(misafir("Serdar", "Vural", "20234567823", "E", 42, "GRUP-15", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Özge",   "Vural", "20234567824", "K", 40, "GRUP-15", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Arda",   "Vural", "20234567825", "E", 15, "GRUP-15", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Ela",    "Vural", "20234567826", "K", 13, "GRUP-15", "Aile", "4 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 106 — 2 Kişilik | ÖZEL-1 Balayı (Deniz Manzaralı)
    # ----------------------------------------------------------------
    satirlar.append(oda(106, "2 Kişilik", manzara="Deniz", ozel_not="VIP - Balayı Çifti"))
    satirlar.append(misafir("Erdem", "Akman", "30234567827", "E", 27, "ÖZEL-1", "Evli", "2 Kişilik",
                            uyari="Deniz manzaralı oda talep edildi"))
    satirlar.append(misafir("Melis", "Akman", "30234567828", "K", 25, "ÖZEL-1", "Evli", "2 Kişilik",
                            uyari="Deniz manzaralı oda talep edildi"))

    # ----------------------------------------------------------------
    # ODA 107 — 2 Kişilik | ÖZEL-2 Engelli Misafir (Zemin Kat)
    # ----------------------------------------------------------------
    satirlar.append(oda(107, "2 Kişilik", ozel_not="Engelli Erişimi - Zemin Kat"))
    satirlar.append(misafir("Halil",  "Başak", "40234567829", "E", 55, "ÖZEL-2", "Evli", "2 Kişilik",
                            uyari="Tekerlekli sandalye - zemin kata yakın oda zorunlu"))
    satirlar.append(misafir("Nermin", "Başak", "40234567830", "K", 53, "ÖZEL-2", "Evli", "2 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 108 — 2 Kişilik | ÖZEL-3 Alerjisi Olan Misafir
    # ----------------------------------------------------------------
    satirlar.append(oda(108, "2 Kişilik", ozel_not="Fındık Alerjisi - Özel Temizlik"))
    satirlar.append(misafir("Levent", "Çakır", "50234567831", "E", 32, "ÖZEL-3", "Evli", "2 Kişilik"))
    satirlar.append(misafir("Tülin",  "Çakır", "50234567832", "K", 30, "ÖZEL-3", "Evli", "2 Kişilik",
                            uyari="Ağır fındık alerjisi - oda temizliğinde dikkat"))

    # ----------------------------------------------------------------
    # ODA 291 — 4 Kişilik | PROBLEM-1 (5 erkek → 4+1)
    # ----------------------------------------------------------------
    satirlar.append(oda(291, "4 Kişilik"))
    satirlar.append(misafir("Tarık",  "Erdoğan", "60234567833", "E", 22, "PROBLEM-1", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Uğur",   "Yalçın",  "60234567834", "E", 21, "PROBLEM-1", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Yasin",  "Kara",    "60234567835", "E", 22, "PROBLEM-1", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Furkan", "Mutlu",   "60234567836", "E", 23, "PROBLEM-1", "Arkadaş", "4 Kişilik"))

    # ODA 109 — 2 Kişilik | PROBLEM-1 (kalan 1 kişi)
    satirlar.append(oda(109, "2 Kişilik"))
    satirlar.append(misafir("Hasan", "Sezer", "60234567837", "E", 22, "PROBLEM-1", "Arkadaş", "4 Kişilik",
                            uyari="Gruptan ayrı odada (5 kişi → 4+1 bölünme)"))

    # ----------------------------------------------------------------
    # ODA 110 — 2 Kişilik | PROBLEM-2 Erkekler (Cinsiyet ayrımı)
    # ----------------------------------------------------------------
    satirlar.append(oda(110, "2 Kişilik"))
    satirlar.append(misafir("Mark",  "Johnson", "PASAPORT-US001", "E", 30, "PROBLEM-2", "Arkadaş", "4 Kişilik",
                            uyari="Evli/nişanlı olmayan karışık grup — cinsiyet bazlı ayrıldı"))
    satirlar.append(misafir("David", "Brown",   "PASAPORT-UK003", "E", 29, "PROBLEM-2", "Arkadaş", "4 Kişilik",
                            uyari="Evli/nişanlı olmayan karışık grup — cinsiyet bazlı ayrıldı"))

    # ODA 111 — 2 Kişilik | PROBLEM-2 Kadınlar
    satirlar.append(oda(111, "2 Kişilik"))
    satirlar.append(misafir("Sarah", "Williams", "PASAPORT-US002", "K", 28, "PROBLEM-2", "Arkadaş", "4 Kişilik",
                            uyari="Evli/nişanlı olmayan karışık grup — cinsiyet bazlı ayrıldı"))
    satirlar.append(misafir("Emma",  "Davis",    "PASAPORT-UK004", "K", 27, "PROBLEM-2", "Arkadaş", "4 Kişilik",
                            uyari="Evli/nişanlı olmayan karışık grup — cinsiyet bazlı ayrıldı"))

    # ----------------------------------------------------------------
    # ODA 292 — 4 Kişilik | PROBLEM-3 (3 kişi 4'lük odada)
    # ----------------------------------------------------------------
    satirlar.append(oda(292, "4 Kişilik"))
    satirlar.append(misafir("Özkan",  "Tunç", "70234567838", "E", 26, "PROBLEM-3", "Arkadaş", "4 Kişilik",
                            uyari="3 kişi — 3'lük oda kalmadığı için 4'lük odaya yerleştirildi"))
    satirlar.append(misafir("Gökhan", "Tan",  "70234567839", "E", 27, "PROBLEM-3", "Arkadaş", "4 Kişilik",
                            uyari="3 kişi — 3'lük oda kalmadığı için 4'lük odaya yerleştirildi"))
    satirlar.append(misafir("Bülent", "İnce", "70234567840", "E", 26, "PROBLEM-3", "Arkadaş", "4 Kişilik",
                            uyari="3 kişi — 3'lük oda kalmadığı için 4'lük odaya yerleştirildi"))

    # ----------------------------------------------------------------
    # ODA 112 — 2 Kişilik | PROBLEM-4 Son Dakika Rezervasyon
    # ----------------------------------------------------------------
    satirlar.append(oda(112, "2 Kişilik"))
    satirlar.append(misafir("Yusuf", "Turan", "80234567841", "E", 50, "PROBLEM-4", "Evli", "2 Kişilik",
                            uyari="Son dakika rezervasyon — oda bulundu"))
    satirlar.append(misafir("Zehra", "Turan", "80234567842", "K", 48, "PROBLEM-4", "Evli", "2 Kişilik",
                            uyari="Son dakika rezervasyon — oda bulundu"))

    # ----------------------------------------------------------------
    # ODA 293 — 4 Kişilik | GRUP-16 Özer Ailesi
    # ----------------------------------------------------------------
    satirlar.append(oda(293, "4 Kişilik"))
    satirlar.append(misafir("Turgut", "Özer", "12345678905", "E", 44, "GRUP-16", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Şükran", "Özer", "12345678906", "K", 42, "GRUP-16", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Enes",   "Özer", "12345678907", "E", 16, "GRUP-16", "Aile", "4 Kişilik"))
    satirlar.append(misafir("Sude",   "Özer", "12345678908", "K", 13, "GRUP-16", "Aile", "4 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 207 — 3 Kişilik | GRUP-17 Kılınç Ailesi
    # ----------------------------------------------------------------
    satirlar.append(oda(207, "3 Kişilik"))
    satirlar.append(misafir("Fırat", "Kılınç", "12345678909", "E", 39, "GRUP-17", "Aile", "3 Kişilik"))
    satirlar.append(misafir("Serap", "Kılınç", "12345678910", "K", 37, "GRUP-17", "Aile", "3 Kişilik"))
    satirlar.append(misafir("Alp",   "Kılınç", "12345678911", "E", 11, "GRUP-17", "Aile", "3 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 113 — 2 Kişilik | GRUP-18 Erbay (Evli çift)
    # ----------------------------------------------------------------
    satirlar.append(oda(113, "2 Kişilik"))
    satirlar.append(misafir("Cengiz", "Erbay", "12345678912", "E", 34, "GRUP-18", "Evli", "2 Kişilik"))
    satirlar.append(misafir("Meltem", "Erbay", "12345678913", "K", 32, "GRUP-18", "Evli", "2 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 114 — 2 Kişilik | GRUP-19 Demirci (Evli çift)
    # ----------------------------------------------------------------
    satirlar.append(oda(114, "2 Kişilik"))
    satirlar.append(misafir("Raşit", "Demirci", "12345678914", "E", 29, "GRUP-19", "Evli", "2 Kişilik"))
    satirlar.append(misafir("Gamze", "Demirci", "12345678915", "K", 28, "GRUP-19", "Evli", "2 Kişilik"))

    # ----------------------------------------------------------------
    # ODA 294 — 4 Kişilik | GRUP-20 Erkek Arkadaş
    # ----------------------------------------------------------------
    satirlar.append(oda(294, "4 Kişilik"))
    satirlar.append(misafir("Ufuk",    "Çetin",  "12345678916", "E", 24, "GRUP-20", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Özgür",   "Işık",   "12345678917", "E", 25, "GRUP-20", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Mert",    "Arslan",  "12345678918", "E", 24, "GRUP-20", "Arkadaş", "4 Kişilik"))
    satirlar.append(misafir("Batuhan", "Yurt",   "12345678919", "E", 23, "GRUP-20", "Arkadaş", "4 Kişilik"))

    return satirlar


def create_excel(data: dict, output_path: str) -> None:
    """
    Turizm ve Konaklama sektörü için kısıt bazlı oda yerleştirme planını
    profesyonel formatta Excel dosyasına dönüştürür.

    Args:
        data: {"satirlar": [...]} formatında yerleştirme verisi
        output_path: Çıktı dosyasının kaydedileceği yol
    """
    if isinstance(data, dict):
        rows = data.get("satirlar", [])
    else:
        rows = data if data else []

    # Kullanıcıdan veri gelmemişse varsayılan demo veriyi kullan
    if not rows:
        rows = _get_default_data()

    wb = Workbook()
    ws = wb.active
    ws.title = "Oda Yerlestirme Plani"

    # --- Stil Tanımları ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", fill_type="solid")

    room_header_fill = PatternFill(start_color="ADD8E6", fill_type="solid")
    warning_fill = PatternFill(start_color="FF0000", fill_type="solid")
    warning_font = Font(color="FFFFFF", bold=True)

    male_fill   = PatternFill(start_color="AED6F1", fill_type="solid")
    female_fill = PatternFill(start_color="F9E7EF", fill_type="solid")
    waitlist_fill = PatternFill(start_color="F7DC6F", fill_type="solid")
    zebra_fill  = PatternFill(start_color="F2F2F2", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_border = Border(
        left=Side(style="medium", color="999999"),
        right=Side(style="medium", color="999999"),
        top=Side(style="medium", color="999999"),
        bottom=Side(style="medium", color="999999"),
    )

    # --- Başlık Satırı ---
    headers = [
        "Satır Tipi", "Oda No", "Oda Tipi", "Manzara", "Özel Not (Oda)",
        "İsim", "Soyisim", "TC/Pasaport", "Cinsiyet", "Yaş",
        "Grup ID", "İlişki Tipi", "Oda Tercihi", "Durum", "Uyarı Açıklama",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    # --- Veri İşleme ---
    current_row = 2
    for item in rows:
        row_type = item.get("Satır Tipi")

        if row_type == "ODA_BAŞLIK" and current_row > 2:
            current_row += 1  # Oda blokları arasına boşluk

        for col_idx, header in enumerate(headers, start=1):
            val = item.get(header)

            # TC Maskeleme
            if header == "TC/Pasaport" and val and not str(val).startswith("PASAPORT"):
                s_val = str(val)
                if len(s_val) > 4:
                    val = "*" * (len(s_val) - 4) + s_val[-4:]

            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = thin_border

            if header == "Oda No":
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            else:
                cell.alignment = Alignment(horizontal="left")

            if row_type == "ODA_BAŞLIK":
                cell.fill = room_header_fill
                cell.font = Font(bold=True)
            elif row_type == "UYARI":
                cell.fill = warning_fill
                cell.font = warning_font
            else:
                if current_row % 2 == 0:
                    cell.fill = zebra_fill

                if header == "Cinsiyet":
                    if val == "E":
                        cell.fill = male_fill
                    elif val == "K":
                        cell.fill = female_fill

                if header == "Durum" and val == "Bekleme Listesi":
                    cell.fill = waitlist_fill

        current_row += 1

    # --- Toplam Satırı ---
    sum_row = current_row
    ws.cell(row=sum_row, column=1, value="TOPLAM / ÖZET").font = Font(bold=True)
    ws.cell(row=sum_row, column=10, value=f"=SUM(J2:J{sum_row - 1})").font = Font(bold=True)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=sum_row, column=col_idx)
        cell.fill = PatternFill(start_color="D6EAF8", fill_type="solid")
        cell.border = Border(top=Side(style="medium"))

    # --- Sayfa Ayarları ---
    ws.freeze_panes = "A2"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # --- Sütun Genişlikleri ---
    column_widths = [15, 10, 15, 12, 35, 15, 15, 18, 10, 8, 12, 15, 15, 18, 60]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
