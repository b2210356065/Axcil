def create_excel(data: dict, output_path: str) -> None:
    """
    Turizm ajansı oda yerleştirme sistemi - 1000 kişilik grup, 230 oda
    data parametresi: Müşteri verisi (boş veya None ise mock veri üretilir)
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # ========================================================================
    # AŞAMA 1: ODA ENVANTERİ OLUŞTURMA
    # ========================================================================
    # 100x2kişilik(200) + 80x3kişilik(240) + 50x4kişilik(200) = 640 kişi
    rooms = []
    room_no = 101
    
    # 2 Kişilik Odalar (100 adet)
    for _ in range(100):
        rooms.append({
            'no': room_no,
            'tip': '2 Kişilik',
            'kapasite': 2,
            'manzara': 'Standart',
            'ozel_not': None,
            'misafirler': []
        })
        room_no = room_no + 1
    
    # 3 Kişilik Odalar (80 adet)
    for _ in range(80):
        rooms.append({
            'no': room_no,
            'tip': '3 Kişilik',
            'kapasite': 3,
            'manzara': 'Standart',
            'ozel_not': None,
            'misafirler': []
        })
        room_no = room_no + 1
    
    # 4 Kişilik Odalar (50 adet)
    for _ in range(50):
        rooms.append({
            'no': room_no,
            'tip': '4 Kişilik',
            'kapasite': 4,
            'manzara': 'Standart',
            'ozel_not': None,
            'misafirler': []
        })
        room_no = room_no + 1
    
    # ========================================================================
    # AŞAMA 2: MİSAFİR VERİSİ OLUŞTURMA
    # ========================================================================
    guests = []
    
    # data boş veya None kontrolü
    data_list = None
    if data is not None:
        if isinstance(data, dict):
            if 'guests' in data:
                data_list = data['guests']
            elif 'misafirler' in data:
                data_list = data['misafirler']
            elif len(data) > 0:
                data_list = None
        elif isinstance(data, list):
            data_list = data
    
    if data_list is None or len(data_list) == 0:
        # Mock veri üret (1000 kişi)
        
        # SENARYO 1: 4 Kişilik Aile - Farklı Cinsiyetler (İlişki='Aile' ZORUNLU)
        guests.append({'isim': 'Mehmet', 'soyisim': 'YILMAZ', 'tc': '12345678901', 'cinsiyet': 'E', 'yas': 45, 'grup_id': 'GRUP-1', 'iliski': 'Aile', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Ayşe', 'soyisim': 'YILMAZ', 'tc': '12345678902', 'cinsiyet': 'K', 'yas': 42, 'grup_id': 'GRUP-1', 'iliski': 'Aile', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Can', 'soyisim': 'YILMAZ', 'tc': '12345678903', 'cinsiyet': 'E', 'yas': 18, 'grup_id': 'GRUP-1', 'iliski': 'Aile', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Elif', 'soyisim': 'YILMAZ', 'tc': '12345678904', 'cinsiyet': 'K', 'yas': 15, 'grup_id': 'GRUP-1', 'iliski': 'Aile', 'tercih': '4 Kişilik', 'ozel': None})
        
        # SENARYO 2: Arkadaş Grubu - Cinsiyet Uyumsuzluğu (2E+2K AYRI odalara)
        guests.append({'isim': 'Ahmet', 'soyisim': 'DEMİR', 'tc': '23456789001', 'cinsiyet': 'E', 'yas': 30, 'grup_id': 'GRUP-5', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Zeynep', 'soyisim': 'KAYA', 'tc': '23456789002', 'cinsiyet': 'K', 'yas': 28, 'grup_id': 'GRUP-5', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Burak', 'soyisim': 'ÇELİK', 'tc': '23456789003', 'cinsiyet': 'E', 'yas': 29, 'grup_id': 'GRUP-5', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Selin', 'soyisim': 'AKIN', 'tc': '23456789004', 'cinsiyet': 'K', 'yas': 27, 'grup_id': 'GRUP-5', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        
        # SENARYO 3: 5 Kişilik Grup - Dağıtılacak (4+1 veya 3+2)
        guests.append({'isim': 'Ali', 'soyisim': 'VURAL', 'tc': '34567890001', 'cinsiyet': 'E', 'yas': 25, 'grup_id': 'GRUP-10', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Veli', 'soyisim': 'ÖZKAN', 'tc': '34567890002', 'cinsiyet': 'E', 'yas': 26, 'grup_id': 'GRUP-10', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Hasan', 'soyisim': 'KURT', 'tc': '34567890003', 'cinsiyet': 'E', 'yas': 27, 'grup_id': 'GRUP-10', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Hüseyin', 'soyisim': 'ASLAN', 'tc': '34567890004', 'cinsiyet': 'E', 'yas': 28, 'grup_id': 'GRUP-10', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Murat', 'soyisim': 'YILDIZ', 'tc': '34567890005', 'cinsiyet': 'E', 'yas': 29, 'grup_id': 'GRUP-10', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        
        # SENARYO 4: Tek Kişi Eşleştirme (2K aynı odada, 1E ayrı)
        guests.append({'isim': 'Fatma', 'soyisim': 'ÖZGÜR', 'tc': '45678900001', 'cinsiyet': 'K', 'yas': 35, 'grup_id': 'GRUP-20', 'iliski': 'Arkadaş', 'tercih': '2 Kişilik', 'ozel': None})
        guests.append({'isim': 'Ayşegül', 'soyisim': 'TOPRAK', 'tc': '45678900002', 'cinsiyet': 'K', 'yas': 40, 'grup_id': 'GRUP-21', 'iliski': 'Arkadaş', 'tercih': '2 Kişilik', 'ozel': None})
        guests.append({'isim': 'Kemal', 'soyisim': 'AKAR', 'tc': '45678900003', 'cinsiyet': 'E', 'yas': 38, 'grup_id': 'GRUP-22', 'iliski': 'Arkadaş', 'tercih': '2 Kişilik', 'ozel': None})
        
        # SENARYO 5: Özel İstek - Deniz + Engelli (Evli çift)
        guests.append({'isim': 'Cem', 'soyisim': 'GÜNEŞ', 'tc': '56789000001', 'cinsiyet': 'E', 'yas': 30, 'grup_id': 'GRUP-50', 'iliski': 'Evli', 'tercih': '2 Kişilik', 'ozel': 'Deniz manzaralı, Engelli erişimi'})
        guests.append({'isim': 'Seda', 'soyisim': 'GÜNEŞ', 'tc': '56789000002', 'cinsiyet': 'K', 'yas': 28, 'grup_id': 'GRUP-50', 'iliski': 'Evli', 'tercih': '2 Kişilik', 'ozel': 'Deniz manzaralı, Engelli erişimi'})
        
        # SENARYO 6: Tercih Dolu - 3 kişi 4'lük ister ama 3'lüğe yerleşir
        guests.append({'isim': 'Osman', 'soyisim': 'POLAT', 'tc': '67890000001', 'cinsiyet': 'E', 'yas': 32, 'grup_id': 'GRUP-30', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Yusuf', 'soyisim': 'ŞAHIN', 'tc': '67890000002', 'cinsiyet': 'E', 'yas': 33, 'grup_id': 'GRUP-30', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        guests.append({'isim': 'Mustafa', 'soyisim': 'KOÇAK', 'tc': '67890000003', 'cinsiyet': 'E', 'yas': 31, 'grup_id': 'GRUP-30', 'iliski': 'Arkadaş', 'tercih': '4 Kişilik', 'ozel': None})
        
        # Kalan kişiler (Kapasite aşımı testi: 640 sınır)
        tercih_listesi = ['2 Kişilik', '3 Kişilik', '4 Kişilik']
        for idx in range(21, 1000):
            guests.append({
                'isim': 'Misafir' + str(idx),
                'soyisim': 'SOYAD' + str(idx),
                'tc': str(10000000000 + idx),
                'cinsiyet': 'E' if idx % 2 == 0 else 'K',
                'yas': 20 + (idx % 50),
                'grup_id': 'GRUP-' + str(100 + (idx // 4)),
                'iliski': 'Arkadaş',
                'tercih': tercih_listesi[idx % 3],
                'ozel': None
            })
    else:
        guests = list(data_list)
    
    # ========================================================================
    # AŞAMA 3: YERLEŞTİRME ALGORİTMASI
    # ========================================================================
    waiting_list = []
    
    def can_room_together(g1, g2):
        """İki misafir aynı odada kalabilir mi?"""
        if g1['grup_id'] == g2['grup_id']:
            if g1['cinsiyet'] != g2['cinsiyet']:
                iliski = g1.get('iliski', 'Arkadaş')
                if iliski in ['Aile', 'Evli', 'Nişanlı', 'Birlikte Kalabilir']:
                    return True
                return False
            return True
        return g1['cinsiyet'] == g2['cinsiyet']
    
    def can_add_to_room(guest, room):
        """Misafir bu odaya eklenebilir mi?"""
        if len(room['misafirler']) == 0:
            return True
        for existing in room['misafirler']:
            if not can_room_together(guest, existing):
                return False
        return True
    
    def get_capacity_order(preferred):
        """Tercih sırasına göre kapasite listesi"""
        if preferred == '4 Kişilik':
            return [4, 3, 2]
        elif preferred == '3 Kişilik':
            return [3, 4, 2]
        else:
            return [2, 3, 4]
    
    # Grupları oluştur
    groups_dict = {}
    for g in guests:
        gid = g['grup_id']
        if gid not in groups_dict:
            groups_dict[gid] = []
        groups_dict[gid].append(g)
    
    # Büyükten küçüğe sırala (optimizasyon)
    sorted_groups = sorted(groups_dict.items(), key=lambda x: len(x[1]), reverse=True)
    
    # GRUP YERLEŞTİRME
    for gid, members in sorted_groups:
        if len(members) == 1:
            continue
        
        group_size = len(members)
        preferred = members[0].get('tercih', '2 Kişilik')
        cap_order = get_capacity_order(preferred)
        
        placed = False
        for cap in cap_order:
            if placed:
                break
            for room in rooms:
                if room['kapasite'] == cap and len(room['misafirler']) == 0:
                    if group_size <= cap:
                        can_place_all = True
                        for m in members:
                            if not can_add_to_room(m, room):
                                can_place_all = False
                                break
                        
                        if can_place_all:
                            for m in members:
                                room['misafirler'].append(m)
                                ozel = m.get('ozel', None)
                                if ozel is not None:
                                    if 'Deniz' in str(ozel):
                                        room['manzara'] = 'Deniz'
                                    if 'Engelli' in str(ozel):
                                        room['ozel_not'] = 'Engelli Erişimi Gerekli'
                            placed = True
                            break
        
        if not placed:
            for m in members:
                m_placed = False
                preferred_m = m.get('tercih', '2 Kişilik')
                cap_order_m = get_capacity_order(preferred_m)
                for cap in cap_order_m:
                    if m_placed:
                        break
                    for room in rooms:
                        if room['kapasite'] == cap and len(room['misafirler']) < room['kapasite']:
                            if can_add_to_room(m, room):
                                room['misafirler'].append(m)
                                ozel = m.get('ozel', None)
                                if ozel is not None:
                                    if 'Deniz' in str(ozel):
                                        room['manzara'] = 'Deniz'
                                    if 'Engelli' in str(ozel):
                                        room['ozel_not'] = 'Engelli Erişimi Gerekli'
                                m_placed = True
                                break
                if not m_placed:
                    waiting_list.append(m)
    
    # TEK KİŞİ YERLEŞTİRME
    for gid, members in sorted_groups:
        if len(members) == 1:
            guest = members[0]
            preferred = guest.get('tercih', '2 Kişilik')
            cap_order = get_capacity_order(preferred)
            
            placed = False
            for cap in cap_order:
                if placed:
                    break
                for room in rooms:
                    if room['kapasite'] == cap and len(room['misafirler']) < room['kapasite']:
                        if can_add_to_room(guest, room):
                            room['misafirler'].append(guest)
                            ozel = guest.get('ozel', None)
                            if ozel is not None:
                                if 'Deniz' in str(ozel):
                                    room['manzara'] = 'Deniz'
                                if 'Engelli' in str(ozel):
                                    room['ozel_not'] = 'Engelli Erişimi Gerekli'
                            placed = True
                            break
            
            if not placed:
                waiting_list.append(guest)
    
    # ========================================================================
    # AŞAMA 4: EXCEL OLUŞTURMA
    # ========================================================================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Oda Yerleşim Planı"
    
    headers = ['Satır Tipi', 'Oda No', 'Oda Tipi', 'Manzara', 'Özel Not (Oda)', 
               'İsim', 'Soyisim', 'TC/Pasaport', 'Cinsiyet', 'Yaş', 'Grup ID', 
               'İlişki Tipi', 'Oda Tercihi', 'Durum', 'Uyarı Açıklama']
    ws.append(headers)
    
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    oda_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    oda_font = Font(bold=True)
    uyari_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    uyari_font = Font(bold=True, color='FFFFFF')
    e_fill = PatternFill(start_color='CCEBFF', end_color='CCEBFF', fill_type='solid')
    k_fill = PatternFill(start_color='FFCCFF', end_color='FFCCFF', fill_type='solid')
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for room in rooms:
        if len(room['misafirler']) == 0:
            continue
        
        ozel_not_val = room['ozel_not'] if room['ozel_not'] is not None else ''
        oda_row = [
            'ODA_BAŞLIK',
            room['no'],
            room['tip'],
            room['manzara'],
            ozel_not_val,
            '', '', '', '', '', '', '', '', '', ''
        ]
        ws.append(oda_row)
        
        current_row = ws.max_row
        for col_idx in range(1, 16):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = oda_fill
            cell.font = oda_font
        
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
        
        for guest in room['misafirler']:
            misafir_row = [
                'MİSAFİR',
                '', '', '', '',
                guest.get('isim', ''),
                guest.get('soyisim', ''),
                guest.get('tc', ''),
                guest.get('cinsiyet', ''),
                guest.get('yas', ''),
                guest.get('grup_id', ''),
                guest.get('iliski', 'Arkadaş'),
                guest.get('tercih', ''),
                'Yerleştirildi',
                ''
            ]
            ws.append(misafir_row)
            
            current_row = ws.max_row
            if guest.get('cinsiyet', '') == 'E':
                ws.cell(row=current_row, column=9).fill = e_fill
            else:
                ws.cell(row=current_row, column=9).fill = k_fill
        
        ws.append([''] * 15)
    
    for guest in waiting_list:
        tercih_val = guest.get('tercih', '2 Kişilik')
        uyari_aciklama = 'Oda bulunamadı - ' + tercih_val + ' ve diğer tüm oda tipleri dolu. Toplam kapasite 640 kişi aşıldı.'
        uyari_row = [
            'UYARI',
            '', '', '', '',
            guest.get('isim', ''),
            guest.get('soyisim', ''),
            guest.get('tc', ''),
            guest.get('cinsiyet', ''),
            guest.get('yas', ''),
            guest.get('grup_id', ''),
            guest.get('iliski', 'Arkadaş'),
            tercih_val,
            'Bekleme Listesi',
            uyari_aciklama
        ]
        ws.append(uyari_row)
        
        current_row = ws.max_row
        for col_idx in range(1, 16):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = uyari_fill
            cell.font = uyari_font
    
    column_widths = [15, 10, 15, 12, 35, 15, 15, 15, 10, 8, 12, 18, 15, 18, 80]
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    for idx in range(len(column_widths)):
        ws.column_dimensions[col_letters[idx]].width = column_widths[idx]
    
    wb.save(output_path)
    return output_path