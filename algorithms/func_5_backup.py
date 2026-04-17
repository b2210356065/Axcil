import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os.path

def create_excel(data: dict, output_path: str) -> None:
    """
    Turizm acentesi oda yerleşim planını profesyonel Excel formatında oluşturur.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Oda_Yerlesim_Plani"

    # Sütun Başlıkları
    headers = [
        "GRUP ADI / ODA NO", 
        "SOYİSİM / ODA TİPİ", 
        "TC-PASAPORT / MANZARA", 
        "CİNSİYET / ÖZEL NOT", 
        "YAŞ", 
        "GRUP ID"
    ]

    # Stiller
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    zebra_even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    zebra_odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    summary_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color="CCCCCC"),
        right=Side(style='thin', color="CCCCCC"),
        top=Side(style='thin', color="CCCCCC"),
        bottom=Side(style='thin', color="CCCCCC")
    )
    
    header_border = Border(
        left=Side(style='medium', color="999999"),
        right=Side(style='medium', color="999999"),
        top=Side(style='medium', color="999999"),
        bottom=Side(style='medium', color="999999")
    )

    # Başlıkları Yaz
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    current_row = 2
    last_room_id = None
    data_rows = data.get("satirlar", [])

    # Verileri Yaz
    for index, row_data in enumerate(data_rows):
        current_room_val = str(row_data.get("GRUP ADI / ODA NO", ""))
        
        # Görsel Boşluk Kontrolü
        if last_room_id is not None and "ODA" in current_room_val and current_room_val != last_room_id:
            current_row += 1

        if "ODA" in current_room_val:
            last_room_id = current_room_val

        for col_num, header in enumerate(headers, 1):
            val = row_data.get(header)
            cell = ws.cell(row=current_row, column=col_num, value=val)
            cell.border = thin_border
            
            if header == "YAŞ":
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "0"
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

            if current_row % 2 == 0:
                cell.fill = zebra_even_fill
            else:
                cell.fill = zebra_odd_fill
        
        current_row += 1

    # Özet Satırları
    summary_start_row = current_row + 1
    summaries = [
        ("Toplam Kişi Sayısı", f"=COUNT({get_column_letter(5)}2:{get_column_letter(5)}{current_row-1})"),
        ("Dolu Oda Sayısı", f"=COUNTIF({get_column_letter(1)}2:{get_column_letter(1)}{current_row-1}, \"*ODA*\")"),
        ("Bekleme Listesindeki Kişi Sayısı", f"=COUNTIF({get_column_letter(1)}2:{get_column_letter(1)}{current_row-1}, \"*UYARI*\")")
    ]

    for label, formula in summaries:
        ws.merge_cells(start_row=summary_start_row, start_column=1, end_row=summary_start_row, end_column=4)
        lbl_cell = ws.cell(row=summary_start_row, column=1, value=label)
        lbl_cell.font = Font(bold=True)
        lbl_cell.alignment = Alignment(horizontal="right")
        lbl_cell.fill = summary_fill
        
        f_cell = ws.cell(row=summary_start_row, column=5, value=formula)
        f_cell.font = Font(bold=True)
        f_cell.fill = summary_fill
        f_cell.border = Border(top=Side(style='medium'))
        summary_start_row += 1

    # Sayfa Ayarları
    ws.freeze_panes = "A2"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Otomatik Sütun Genişliği
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 40)

    wb.save(output_path)
