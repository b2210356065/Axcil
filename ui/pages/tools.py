"""Araclar sayfasi - Coklu arac secimi ve Excel CRUD islemleri."""
import streamlit as st
import tempfile
import os
import json
import datetime
import traceback
from io import BytesIO
from copy import deepcopy

from typing import Any, List

import pandas as pd
from pydantic import BaseModel, Field

from core.database import get_functionalities, get_active_business, get_functionality_by_id
from core.models import TaskType
from core.algorithm_runner import algorithm_exists, run_algorithm
from core.algorithm_generator import generate_algorithm
from core.enrichment import enrich_functionality, confirm_enrichment


# ---------------------------------------------------------------------------
# Pydantic semalari
# ---------------------------------------------------------------------------

class ExcelSheet(BaseModel):
    ad: str = Field(description="Sayfa adi")
    basliklar: List[str] = Field(description="Sutun basliklari")
    veriler: List[List[Any]] = Field(description="Satir verileri")


class ExcelData(BaseModel):
    sheetler: List[ExcelSheet] = Field(description="Excel sayfalari")


class DeleteInstruction(BaseModel):
    key_column: str = Field(description="Satirlari benzersiz tanimlayan sutun adi")
    delete_keys: List[str] = Field(description="Silinecek satirlarin anahtar degerleri (string olarak)")


class EditChange(BaseModel):
    key_value: str = Field(description="Duzenlenecek satirin anahtar degeri")
    changes: dict = Field(description="Degistirilecek sutun adi ve yeni deger ciftleri")


class EditInstruction(BaseModel):
    key_column: str = Field(description="Satirlari benzersiz tanimlayan sutun adi")
    edits: List[EditChange] = Field(description="Duzenlenecek satirlar ve degisiklikler")


class VerifiedExcelData(BaseModel):
    key_column: str = Field(description="Satirlari benzersiz tanimlayan anahtar sutun adi")
    key_values: List[str] = Field(description="Girdi verisinden cikartilan TUM anahtar degerler listesi")
    sheetler: List[ExcelSheet] = Field(description="Excel sayfalari")


# ---------------------------------------------------------------------------
# Sabitler
# ---------------------------------------------------------------------------

TOOLS = [
    {"id": "text",  "icon": "T",  "name": "Metin",  "desc": "Serbest metin, CSV, JSON veya e-posta"},
    {"id": "image", "icon": "I",  "name": "Gorsel", "desc": "Fotograflardan veri cikarma"},
    {"id": "pdf",   "icon": "P",  "name": "PDF",    "desc": "PDF belgelerden veri cikarma"},
    {"id": "voice", "icon": "S",  "name": "Ses",    "desc": "Sesli notlari donusturme"},
    {"id": "form",  "icon": "F",  "name": "Form",   "desc": "Yapilandirilmis form girisi"},
    {"id": "excel", "icon": "E",  "name": "Excel",  "desc": "Mevcut Excel dosyasindan veri alma"},
]

TOOL_LOOKUP = {t["id"]: t for t in TOOLS}

MODE_LABELS = {
    "create": "Yeni Olustur",
    "append": "Mevcut Dosyaya Ekle",
    "delete": "Mevcut Dosyadan Sil",
    "edit":   "Mevcut Dosyayi Duzenle",
}


# ---------------------------------------------------------------------------
# Yardimci fonksiyonlar
# ---------------------------------------------------------------------------

def _build_job_context(func) -> str:
    """Is tanimindan prompt baglami olustur."""
    if func:
        system_prompt = func.get("system_prompt", "")
        return f"""
IS TANIMI: {func['name']}
{func['description']}

{system_prompt}
"""
    return "Bu veri Excel formatina donusturulecek.\nVeriyi analiz et ve uygun Excel yapisi belirle.\n"


def _read_existing_excel(uploaded_file):
    """Yuklenen Excel dosyasini oku. Dict listesi doner."""
    try:
        from openpyxl import load_workbook

        file_bytes = uploaded_file.read()
        uploaded_file.seek(0)

        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        sheets = []
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            headers = []
            rows_data = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                row_list = [cell if cell is not None else "" for cell in row]
                if row_idx == 0:
                    headers = [str(h) for h in row_list]
                else:
                    rows_data.append(row_list)
            if headers:
                sheets.append({
                    "ad": ws_name,
                    "basliklar": headers,
                    "veriler": rows_data,
                })
        return {"sheetler": sheets, "raw_bytes": file_bytes}
    except Exception as e:
        st.error(f"Excel okuma hatasi: {e}")
        return None


def _existing_data_to_text(existing_data, max_rows=50):
    """Mevcut Excel verisini prompt icin metin formatina cevir."""
    parts = []
    for sheet in existing_data.get("sheetler", []):
        parts.append(f"--- Sayfa: {sheet['ad']} ---")
        parts.append(f"Basliklar: {json.dumps(sheet['basliklar'], ensure_ascii=False)}")
        sample = sheet["veriler"][:max_rows]
        parts.append(f"Veri ({len(sheet['veriler'])} satir, ilk {len(sample)} gosteriliyor):")
        for row in sample:
            parts.append(f"  {json.dumps([str(v) for v in row], ensure_ascii=False)}")
    return "\n".join(parts)


def _extract_pdf_text(uploaded_file):
    """PDF dosyasindan metin cikar. Basarisizsa None doner."""
    try:
        import fitz
        pdf_bytes = uploaded_file.read()
        uploaded_file.seek(0)
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        pages = []
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            pages.append(f"--- Sayfa {page_num + 1} ---\n{page.get_text()}")
        doc.close()
        text = "\n\n".join(pages)
        return text if text.strip() else None
    except Exception:
        return None


def _transcribe_audio(uploaded_file):
    """Ses dosyasini Whisper ile transkribe et. Basarisizsa None doner."""
    try:
        import openai as openai_pkg
        openai_key = st.session_state.router.config.openai_key
        if not openai_key:
            return None
        client = openai_pkg.OpenAI(api_key=openai_key)
        audio_bytes = uploaded_file.read()
        uploaded_file.seek(0)
        ext = uploaded_file.name.rsplit(".", 1)[-1].lower() if "." in uploaded_file.name else "mp3"
        with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp:
            tmp.write(audio_bytes)
            tmp_path = tmp.name
        try:
            with open(tmp_path, "rb") as af:
                resp = client.audio.transcriptions.create(model="whisper-1", file=af, language="tr")
            return resp.text
        finally:
            os.unlink(tmp_path)
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Excel CRUD algoritmalari (AI kullanmaz, dogrudan islem yapar)
# ---------------------------------------------------------------------------

def _append_to_excel(existing_bytes, new_sheets_data):
    """Mevcut Excel'e yeni satirlar ekle. Mevcut veri korunur."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(BytesIO(existing_bytes))

    for new_sheet in new_sheets_data:
        new_headers = new_sheet["basliklar"]
        new_rows = new_sheet["veriler"]
        target_name = new_sheet.get("ad", wb.sheetnames[0])

        # Hedef sayfayi bul veya ilk sayfaya ekle
        ws = None
        for ws_name in wb.sheetnames:
            if ws_name.lower().strip() == target_name.lower().strip():
                ws = wb[ws_name]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        # Mevcut basliklari oku
        existing_headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(1, col).value
            existing_headers.append(str(val) if val else "")

        # Sutun esleme: yeni baslik -> mevcut sutun indeksi
        col_map = {}
        for new_idx, new_h in enumerate(new_headers):
            for exist_idx, exist_h in enumerate(existing_headers):
                if new_h.lower().strip() == exist_h.lower().strip():
                    col_map[new_idx] = exist_idx + 1  # openpyxl 1-indexed
                    break

        # Eslenmeyenler icin yeni sutunlar ekle
        next_col = ws.max_column + 1
        for new_idx, new_h in enumerate(new_headers):
            if new_idx not in col_map:
                col_map[new_idx] = next_col
                header_cell = ws.cell(1, next_col, value=new_h)
                header_cell.font = Font(bold=True, color="FFFFFF")
                header_cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                header_cell.alignment = Alignment(horizontal="center")
                next_col += 1

        # Yeni satirlari ekle
        start_row = ws.max_row + 1
        for row_data in new_rows:
            for new_idx, value in enumerate(row_data):
                if new_idx in col_map:
                    ws.cell(start_row, col_map[new_idx], value=value)
            start_row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _delete_from_excel(existing_bytes, key_column, delete_keys):
    """Mevcut Excel'den anahtar sutuna gore satirlari sil."""
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(existing_bytes))
    total_deleted = 0

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        # Anahtar sutunu bul
        key_col_idx = None
        for col in range(1, ws.max_column + 1):
            if str(ws.cell(1, col).value).lower().strip() == key_column.lower().strip():
                key_col_idx = col
                break
        if key_col_idx is None:
            continue

        # Silinecek satirlari bul (asagidan yukariya)
        delete_keys_str = [str(k).strip() for k in delete_keys]
        rows_to_delete = []
        for row in range(ws.max_row, 1, -1):
            cell_val = str(ws.cell(row, key_col_idx).value).strip()
            if cell_val in delete_keys_str:
                rows_to_delete.append(row)

        for row in rows_to_delete:
            ws.delete_rows(row)
            total_deleted += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), total_deleted


def _edit_excel(existing_bytes, key_column, edits):
    """Mevcut Excel'de anahtar sutuna gore hucreleri duzenle."""
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(existing_bytes))
    total_edited = 0

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        headers = []
        for col in range(1, ws.max_column + 1):
            headers.append(str(ws.cell(1, col).value).lower().strip())

        # Anahtar sutunu bul
        key_lower = key_column.lower().strip()
        key_col_idx = None
        for idx, h in enumerate(headers):
            if h == key_lower:
                key_col_idx = idx + 1
                break
        if key_col_idx is None:
            continue

        for edit_item in edits:
            key_val = str(edit_item.get("key_value", "")).strip()
            changes = edit_item.get("changes", {})

            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row, key_col_idx).value).strip() == key_val:
                    for col_name, new_val in changes.items():
                        col_lower = col_name.lower().strip()
                        for cidx, h in enumerate(headers):
                            if h == col_lower:
                                ws.cell(row, cidx + 1, value=new_val)
                                total_edited += 1
                                break
                    break  # Her anahtar icin tek satir

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), total_edited


# ---------------------------------------------------------------------------
# Excel olusturma ve indirme
# ---------------------------------------------------------------------------

def _create_excel_bytes(data):
    """ExcelData dict'den xlsx bytes olustur."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    sheets = data.get("sheetler", [])
    if not sheets:
        # Bos sheetler listesi — varsayilan bos sayfa birak
        ws = wb.active
        ws.title = "Bos"
        ws.cell(row=1, column=1, value="Veri bulunamadi")
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    wb.remove(wb.active)

    for sheet_data in sheets:
        ws = wb.create_sheet(str(sheet_data["ad"])[:31])
        headers = sheet_data["basliklar"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row_data in enumerate(sheet_data["veriler"], 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    if wb.worksheets:
        wb.active = wb.worksheets[0]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _show_download_button(excel_bytes, prefix="excel"):
    """Excel indirme butonu goster."""
    filename = f"{prefix}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        label="Excel Indir",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )


# ---------------------------------------------------------------------------
# AI cagri fonksiyonlari
# ---------------------------------------------------------------------------

def _call_ai_extract(prompt, schema_class, image_data=None, mime_type="image/jpeg"):
    """AI ile veri cikartma. AIResponse doner."""
    from core.debug_logger import AIDebugLogger
    dbg = AIDebugLogger("tools_call_ai_extract", provider="tools_page", model="auto")

    router = st.session_state.router
    model = router.select_model(TaskType.EXTRACTION)
    st.info(f"Model: {model.provider_name} - {model.model_name}")

    dbg.log_stage("model_selected", {
        "provider": model.provider_name,
        "model": model.model_name,
        "schema_class": schema_class.__name__ if hasattr(schema_class, '__name__') else str(schema_class),
        "prompt_length": len(prompt),
        "prompt_preview": prompt[:500],
        "has_image": image_data is not None,
        "image_size": len(image_data) if image_data else 0,
    })

    try:
        if image_data:
            result = model.extract(prompt, schema_class, image_data=image_data, mime_type=mime_type)
        else:
            result = model.extract(prompt, schema_class)

        dbg.log_stage("extract_result", {
            "success": True,
            "content_length": len(result.content) if result.content else 0,
            "structured_data": result.structured_data,
            "tokens": {"input": result.input_tokens, "output": result.output_tokens},
            "cost_usd": result.cost_usd,
        })
        dbg.finish(success=True, result=result.structured_data)
        return result

    except Exception as e:
        dbg.log_error(e, context="tools_call_ai_extract")
        dbg.finish(success=False)
        raise


# ---------------------------------------------------------------------------
# Prompt olusturma
# ---------------------------------------------------------------------------

def _build_create_prompt(job_context, combined_input):
    """Yeni Excel olusturma promptu — 3 adimli key dogrulama ile."""
    return f"""{job_context}

GIRDI VERISI:
{combined_input}

GOREV (3 ADIMLI — HER ADIMI SIRASI ILE UYGULA):

ADIM 1 — ANAHTAR CIKARMA:
Girdi verisindeki her bir kaydi benzersiz tanimlayan bir ANAHTAR SUTUN belirle.
(Ornegin: isim, fatura no, urun kodu, ogrenci no, musteri adi, vb.)
Bu sutundaki TUM degerleri tek tek say ve listele. HICBIRINI ATLAMA.
Girdide 50 kayit varsa 50 anahtar deger olmali, 100 kayit varsa 100 olmali.

ADIM 2 — TABLO OLUSTURMA:
Tum veriyi yapilandirilmis Excel tablosuna donustur.
Uygun sutun basliklari belirle ve satirlari olustur.
Her bir anahtar degere karsilik gelen tum bilgileri ilgili satira yaz.

ADIM 3 — KENDINI DOGRULA:
Tabloyu olusturduktan sonra ADIM 1'deki listedeki HER anahtar degerin
tabloda bulundugundan emin ol.
- Tablodaki anahtar sutun degerlerini ADIM 1 listesi ile TEKER TEKER karsilastir.
- Eksik satir varsa tabloya EKLE.
- Tablodaki satir sayisi >= ADIM 1'deki anahtar deger sayisi olmali.

CIKTI FORMATI (bu yapiya KESINLIKLE uy):
{{
    "key_column": "anahtar sutun adi",
    "key_values": ["anahtar1", "anahtar2", "anahtar3"],
    "sheetler": [{{
        "ad": "Sayfa1",
        "basliklar": ["Sutun1", "Sutun2"],
        "veriler": [["deger1", "deger2"]]
    }}]
}}

KRITIK KURALLAR:
- key_values listesi girdideki TUM benzersiz kayitlari icermeli
- sheetler icindeki tabloda key_values'daki HER deger bir satirda bulunmali
- Girdide olmayan hayali veri EKLEME
- Hicbir kaydi ATLAMA

SADECE JSON DONDUR, BASKA BIR SEY YAZMA!
"""


def _build_append_prompt(job_context, combined_input, existing_data):
    """Mevcut Excel'e ekleme icin veri formatlama promptu."""
    existing_text = _existing_data_to_text(existing_data)
    existing_headers = existing_data["sheetler"][0]["basliklar"] if existing_data.get("sheetler") else []

    return f"""{job_context}

MEVCUT EXCEL YAPISI:
{existing_text}

MEVCUT BASLIKLAR: {json.dumps(existing_headers, ensure_ascii=False)}

YENI EKLENMEK ISTENEN VERI:
{combined_input}

GOREV:
1. Yeni veriyi analiz et
2. Mevcut Excel'in baslik yapisina KESINLIKLE uyumlu olacak sekilde formatla
3. SADECE yeni eklenecek satirlari dondur (mevcut verileri EKLEME)
4. Basliklar mevcut Excel ile AYNI olmali

CIKTI FORMATI:
{{"sheetler": [{{"ad": "{existing_data['sheetler'][0]['ad'] if existing_data.get('sheetler') else 'Sayfa1'}", "basliklar": {json.dumps(existing_headers, ensure_ascii=False)}, "veriler": [["yeni_deger1", "yeni_deger2"]]}}]}}

ONEMLI:
- Basliklar mevcut dosyadaki ile TAMAMEN AYNI olmali
- Sadece YENI satirlari dondur, mevcut verileri tekrarlama
- Veri tiplerini mevcut verilerle tutarli tut (sayi ise sayi, tarih ise tarih)

SADECE JSON DONDUR!
"""


def _build_delete_prompt(job_context, combined_input, existing_data):
    """Silme islemi icin anahtar sutun ve degerler promptu."""
    existing_text = _existing_data_to_text(existing_data)

    return f"""{job_context}

MEVCUT EXCEL VERISI:
{existing_text}

KULLANICININ SILME TALEBI:
{combined_input}

GOREV:
1. Kullanicinin hangi satirlari silmek istedigini anla
2. Satirlari benzersiz olarak tanimlayan bir anahtar sutun belirle
   (Ornegin: Fatura No, TC Kimlik, Siparis ID, vb. — en benzersiz olan sutun)
3. Silinmesi gereken satirlarin anahtar degerlerini listele

CIKTI FORMATI (bu yapiya kesinlikle uy):
{{"key_column": "sutun_adi", "delete_keys": ["deger1", "deger2", "deger3"]}}

KURALLAR:
- key_column mevcut basliklardan BIRI olmali
- delete_keys icindeki degerler mevcut veride GERCEKTEN BULUNAN degerler olmali
- Kullanicinin belirtmedigi satirlari SILME
- Eger belirsizlik varsa, en az sayida satir sec (ihtiyatli ol)

SADECE JSON DONDUR!
"""


def _build_edit_prompt(job_context, combined_input, existing_data):
    """Duzenleme islemi icin anahtar sutun ve degisiklik promptu."""
    existing_text = _existing_data_to_text(existing_data)

    return f"""{job_context}

MEVCUT EXCEL VERISI:
{existing_text}

KULLANICININ DUZENLEME TALEBI:
{combined_input}

GOREV:
1. Kullanicinin hangi degerleri degistirmek istedigini anla
2. Satirlari benzersiz tanimlayan bir anahtar sutun belirle
3. Her degisiklik icin: anahtar deger + degisen sutun-deger ciftleri belirle

CIKTI FORMATI (bu yapiya kesinlikle uy):
{{"key_column": "sutun_adi", "edits": [{{"key_value": "satir_anahtar", "changes": {{"degisen_sutun": "yeni_deger"}}}}]}}

KURALLAR:
- key_column mevcut basliklardan BIRI olmali
- key_value mevcut veride GERCEKTEN BULUNAN bir deger olmali
- changes icindeki sutun adlari mevcut basliklardan olmali
- SADECE kullanicinin belirttigi degisiklikleri yap, fazladan degisiklik EKLEME
- Belirtilmeyen sutunlari DEGISTIRME

SADECE JSON DONDUR!
"""


def _build_reorganize_prompt(job_context, current_data, key_list=None):
    """Bastan duzenleme promptu. Key listesi ile veri butunlugu kontrolu."""
    data_text = _existing_data_to_text(current_data)
    row_count = sum(len(s["veriler"]) for s in current_data.get("sheetler", []))

    # Key listesini mevcut veriden cikar (verilmediyse)
    if key_list is None:
        key_list = _extract_key_list_from_data(current_data)

    key_info = ""
    if key_list:
        key_info = f"""
ANAHTAR SUTUN: {key_list['key_column']}
ANAHTAR DEGERLER ({len(key_list['key_values'])} adet):
{json.dumps(key_list['key_values'], ensure_ascii=False)}

Bu listedeki HER deger ciktida da bulunmali!
"""

    return f"""{job_context}

MEVCUT EXCEL VERISI:
{data_text}

TOPLAM SATIR SAYISI: {row_count}
{key_info}

GOREV (3 ADIMLI):

ADIM 1 — ANAHTAR KONTROL:
Yukaridaki anahtar degerlerin tamamini not al.

ADIM 2 — YENIDEN DUZENLEME:
Is tanimina gore verileri yeniden duzenle (siralama, kategorilendirme, gruplama).

ADIM 3 — DOGRULAMA:
Duzenleme sonrasi anahtar listesindeki HER degerin ciktida bulundugunu dogrula.
Eksik varsa EKLE.

KRITIK KURALLAR:
1. HICBIR VERIYI SILME — ciktidaki satir sayisi TAM OLARAK {row_count} olmali
2. HAYALI VERI EKLEME — sadece mevcut verileri kullan
3. MEVCUT DEGERLERI DEGISTIRME — sayilari, isimleri, tarihleri oldugu gibi birak
4. SADECE sunlari yapabilirsin:
   - Satirlarin sirasini degistir (gruplama, kronolojik siralama vb.)
   - Kategori/siniflandirma sutunu varsa is tanimina gore duzelt
   - Bos alanlari bos birak (tahminde bulunma)

CIKTI FORMATI:
{{
    "key_column": "{key_list['key_column'] if key_list else 'anahtar_sutun'}",
    "key_values": {json.dumps(key_list['key_values'] if key_list else [], ensure_ascii=False)},
    "sheetler": [{{"ad": "...", "basliklar": [...], "veriler": [[...]]}}]
}}

SADECE JSON DONDUR!
"""


# ---------------------------------------------------------------------------
# Key dogrulama ve hata yonetimi
# ---------------------------------------------------------------------------

def _extract_key_list_from_data(data):
    """Mevcut Excel verisinden anahtar sutun ve degerleri cikar.
    Ilk sutunu anahtar kabul eder (en basit ve guvenilir yontem)."""
    if not data.get("sheetler"):
        return None
    sheet = data["sheetler"][0]
    if not sheet.get("basliklar") or not sheet.get("veriler"):
        return None
    key_col = sheet["basliklar"][0]
    key_vals = []
    for row in sheet["veriler"]:
        if row and str(row[0]).strip():
            key_vals.append(str(row[0]).strip())
    if not key_vals:
        return None
    return {"key_column": key_col, "key_values": key_vals}


def _verify_keys(ai_data, expected_key_col, expected_keys):
    """AI ciktisindaki tabloda beklenen tum keylerin varligini dogrula.
    Eksik keyleri list olarak doner."""
    if not expected_keys:
        return []

    expected_set = set(str(k).strip().lower() for k in expected_keys)

    # AI ciktisindaki anahtar sutun degerlerini topla
    found_set = set()
    key_col_lower = expected_key_col.lower().strip()

    for sheet in ai_data.get("sheetler", []):
        headers_lower = [str(h).lower().strip() for h in sheet.get("basliklar", [])]
        key_idx = None
        for idx, h in enumerate(headers_lower):
            if h == key_col_lower:
                key_idx = idx
                break
        if key_idx is None:
            continue
        for row in sheet.get("veriler", []):
            if key_idx < len(row) and str(row[key_idx]).strip():
                found_set.add(str(row[key_idx]).strip().lower())

    missing = []
    for k in expected_keys:
        if str(k).strip().lower() not in found_set:
            missing.append(str(k).strip())
    return missing


def _add_failed_entries_to_excel(excel_bytes, key_column, missing_keys):
    """Excel'e 'AI bu verileri eklerken basarisiz oldu' satiri ekle."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active

    # 2 bos satir birak
    start_row = ws.max_row + 2

    # Uyari baslik satiri
    cell = ws.cell(start_row, 1,
                   value=f"UYARI: AI asagidaki verileri eklerken basarisiz oldu (anahtar sutun: {key_column})")
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    # Basligi birlestir
    from openpyxl.utils import get_column_letter
    last_col = ws.max_column
    if last_col > 1:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=last_col)

    # Eksik anahtarlari listele
    for idx, key in enumerate(missing_keys):
        row_num = start_row + 1 + idx
        cell = ws.cell(row_num, 1, value=key)
        cell.font = Font(color="CC0000", italic=True)
        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Giris alanlari
# ---------------------------------------------------------------------------

def _show_tool_input(tool_id, func, prefix):
    """Arac icin giris alani goster. Toplanan veriyi dict olarak doner."""
    if tool_id == "text":
        text = st.text_area(
            "Metni yapistirin",
            height=250,
            placeholder="Veriyi buraya yapistirin...\nSerbest metin, CSV, JSON, e-posta, WhatsApp mesaji vb.",
            key=f"text_input_{prefix}",
        )
        return {"type": "text", "content": text}

    elif tool_id == "image":
        files = st.file_uploader(
            "Fotograf yukleyin",
            type=["jpg", "jpeg", "png", "webp"],
            accept_multiple_files=True,
            key=f"image_input_{prefix}",
        )
        if files:
            preview_cols = st.columns(min(len(files), 4))
            for idx, f in enumerate(files[:4]):
                with preview_cols[idx]:
                    st.image(f, caption=f.name, use_container_width=True)
        return {"type": "image", "files": files or []}

    elif tool_id == "pdf":
        file = st.file_uploader(
            "PDF yukleyin",
            type=["pdf"],
            key=f"pdf_input_{prefix}",
        )
        return {"type": "pdf", "file": file}

    elif tool_id == "voice":
        file = st.file_uploader(
            "Ses dosyasi yukleyin",
            type=["mp3", "wav", "m4a", "ogg"],
            key=f"voice_input_{prefix}",
        )
        if file:
            st.audio(file)
        return {"type": "voice", "file": file}

    elif tool_id == "form":
        return _show_form_input(func, prefix)

    elif tool_id == "excel":
        files = st.file_uploader(
            "Excel dosyasi yukleyin (veri kaynagi olarak)",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key=f"excel_input_{prefix}",
        )
        return {"type": "excel", "files": files or []}

    return {"type": "unknown"}


def _show_form_input(func, prefix):
    """Form araci icin giris alani."""
    input_fields = None
    if func:
        raw = func.get("input_fields")
        if raw:
            if isinstance(raw, str):
                try:
                    input_fields = json.loads(raw)
                except Exception:
                    pass
            elif isinstance(raw, list):
                input_fields = raw

    if not input_fields or not isinstance(input_fields, list) or len(input_fields) == 0:
        # Genel form: sutun basliklarini tanimla
        cols_text = st.text_input(
            "Sutun basliklari (virgul ile ayirin)",
            placeholder="Ornek: Ad, Soyad, Tutar, Tarih",
            key=f"form_cols_{prefix}",
        )
        if not cols_text:
            return {"type": "form", "form_data": []}
        headers = [h.strip() for h in cols_text.split(",") if h.strip()]
        if not headers:
            return {"type": "form", "form_data": []}
        input_fields = [{"name": h, "type": "text"} for h in headers]

    num_rows = st.number_input("Satir sayisi", 1, 100, 1, key=f"form_rows_{prefix}")
    all_rows = []
    for r in range(num_rows):
        if num_rows > 1:
            st.caption(f"Satir {r + 1}")
        row = {}
        cols = st.columns(min(len(input_fields), 4))
        for fi, field in enumerate(input_fields):
            fname = field.get("name", f"Alan_{fi}")
            ftype = field.get("type", "text").lower()
            with cols[fi % len(cols)]:
                if ftype in ("number", "currency", "sayi", "para"):
                    row[fname] = st.number_input(fname, key=f"form_{prefix}_{r}_{fi}")
                elif ftype in ("date", "tarih"):
                    row[fname] = str(st.date_input(fname, key=f"form_{prefix}_{r}_{fi}"))
                else:
                    row[fname] = st.text_input(fname, key=f"form_{prefix}_{r}_{fi}")
        all_rows.append(row)

    if all_rows:
        st.markdown("**Form Onizleme:**")
        st.dataframe(pd.DataFrame(all_rows), use_container_width=True, height=150)

    return {"type": "form", "form_data": all_rows}


# ---------------------------------------------------------------------------
# Girdi birlestirme
# ---------------------------------------------------------------------------

def _combine_inputs(inputs):
    """Tum giris verilerini prompt parcalarina ve gorsel listesine ayir."""
    text_parts = []
    image_list = []  # [(bytes, mime_type)]

    for tool_id, data in inputs.items():
        if data["type"] == "text" and data.get("content"):
            text_parts.append(f"METIN VERISI:\n{data['content']}")

        elif data["type"] == "image" and data.get("files"):
            for f in data["files"]:
                img_bytes = f.read()
                f.seek(0)
                ext = f.name.rsplit(".", 1)[-1].lower() if "." in f.name else "jpeg"
                mime_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp"}
                image_list.append((img_bytes, mime_map.get(ext, "image/jpeg")))
            text_parts.append(f"{len(data['files'])} adet gorsel yuklendi (analiz et).")

        elif data["type"] == "pdf" and data.get("file"):
            pdf_text = _extract_pdf_text(data["file"])
            if pdf_text:
                text_parts.append(f"PDF BELGESI ICERIGI:\n{pdf_text[:50000]}")
            else:
                pdf_bytes = data["file"].read()
                data["file"].seek(0)
                image_list.append((pdf_bytes, "application/pdf"))
                text_parts.append("PDF belgesi gorsel olarak yuklendi.")

        elif data["type"] == "voice" and data.get("file"):
            with st.spinner("Ses dosyasi transkribe ediliyor..."):
                transcript = _transcribe_audio(data["file"])
            if transcript:
                st.success("Transkripsiyon tamamlandi")
                with st.expander("Transkript"):
                    st.write(transcript)
                text_parts.append(f"SES KAYDI TRANSKRIPTI:\n{transcript}")
            else:
                # Gemini'ye ses verisi olarak gonder
                audio_bytes = data["file"].read()
                data["file"].seek(0)
                ext = data["file"].name.rsplit(".", 1)[-1].lower() if "." in data["file"].name else "mp3"
                mime_map = {"mp3": "audio/mpeg", "wav": "audio/wav", "m4a": "audio/mp4", "ogg": "audio/ogg"}
                image_list.append((audio_bytes, mime_map.get(ext, "audio/mpeg")))
                text_parts.append("Ses kaydi yuklendi (dinle ve veriyi cikar).")

        elif data["type"] == "form" and data.get("form_data"):
            rows = data["form_data"]
            if rows and any(any(v not in ("", 0, 0.0) for v in row.values()) for row in rows):
                text_parts.append(f"FORM VERISI:\n{json.dumps(rows, ensure_ascii=False, indent=2)}")

        elif data["type"] == "excel" and data.get("files"):
            for f in data["files"]:
                excel_data = _read_existing_excel(f)
                if excel_data:
                    text_parts.append(f"EXCEL KAYNAK VERISI ({f.name}):\n{_existing_data_to_text(excel_data)}")

    combined_text = "\n\n".join(text_parts)
    return combined_text, image_list


def _has_meaningful_input(inputs):
    """En az bir aractan veri girilmis mi?"""
    for data in inputs.values():
        if data["type"] == "text" and data.get("content", "").strip():
            return True
        if data["type"] == "image" and data.get("files"):
            return True
        if data["type"] == "pdf" and data.get("file"):
            return True
        if data["type"] == "voice" and data.get("file"):
            return True
        if data["type"] == "form" and data.get("form_data"):
            rows = data["form_data"]
            if rows and any(any(v not in ("", 0, 0.0) for v in row.values()) for row in rows):
                return True
        if data["type"] == "excel" and data.get("files"):
            return True
    return False


# ---------------------------------------------------------------------------
# Algoritma tabanli akis yardimcilari
# ---------------------------------------------------------------------------

def _ensure_algorithm_ready(func_id: int, router) -> bool:
    """Algoritma yoksa otomatik olustur (lazy creation). Basarili ise True doner."""
    func = get_functionality_by_id(func_id)
    if not func:
        return False

    # Zenginlestirilmis tanim yoksa otomatik zenginlestir
    if not func.get("enriched_definition"):
        st.info("Is tanimi henuz zenginlestirilmemis. Otomatik zenginlestiriliyor...")
        with st.spinner("AI tanimi zenginlestiriyor..."):
            enrich_result = enrich_functionality(func_id, router)
        if not enrich_result["success"]:
            st.error(f"Zenginlestirme basarisiz: {enrich_result.get('error', '')}")
            return False
        # Otomatik onayla (kullanici araçlar sayfasindan geldi, hizli akis)
        confirm_enrichment(func_id, enrich_result["enriched"], enrich_result.get("attempt_id"))
        st.success("Tanim otomatik zenginlestirildi!")

    # Algoritma olustur
    st.info("Is akisi olusturuluyor (ilk seferde bir defaya mahsus)...")
    with st.spinner("AI is akisi olusturuyor (bu biraz zaman alabilir)..."):
        algo_result = generate_algorithm(func_id, router)
    if algo_result["success"]:
        st.success(f"Is akisi basariyla olusturuldu (v{algo_result.get('version', 1)})")
        return True
    else:
        st.warning(f"Is akisi olusturulamadi: {algo_result.get('error', '')}. Eski akis kullanilacak.")
        return False


def _process_create_with_algorithm(func, func_id, combined_text, image_list, job_context):
    """Algoritma tabanli Excel olusturma: veri cikar + run_algorithm."""
    from ai.prompts.generation import RuntimeExtractionPromptBuilder
    from core.enrichment import _parse_json_response
    from core.debug_logger import AIDebugLogger

    dbg = AIDebugLogger("algo_create_flow", provider="tools_page", model="algorithm")

    # Zenginlestirilmis tanimi al
    enriched_raw = func.get("enriched_definition", "{}")
    try:
        enriched = json.loads(enriched_raw) if isinstance(enriched_raw, str) else enriched_raw
    except (json.JSONDecodeError, TypeError):
        enriched = {}

    dbg.log_stage("enriched_definition", {
        "func_id": func_id,
        "func_name": func.get("name", "?"),
        "enriched_keys": list(enriched.keys()) if isinstance(enriched, dict) else str(type(enriched)),
        "enriched_preview": str(enriched)[:500],
    })

    # RuntimeExtractionPromptBuilder ile veri cikarma promptu olustur
    input_type = "metin"
    if image_list:
        input_type = "gorsel"
    extraction_prompt = RuntimeExtractionPromptBuilder.build_runtime_extraction(
        enriched_definition=enriched,
        input_type=input_type,
    )
    # Kullanici girdisini prompt'a ekle
    full_prompt = f"{extraction_prompt}\n\n<girdi>\n{combined_text}\n</girdi>"

    dbg.log_prompt(full_prompt, extra={
        "input_type": input_type,
        "extraction_prompt_length": len(extraction_prompt),
        "combined_text_length": len(combined_text),
        "image_count": len(image_list),
    })

    # AI ile veri cikar (JSON satirlar formatinda)
    router = st.session_state.router
    model = router.select_model(TaskType.EXTRACTION)
    st.info(f"Veri cikartma: {model.provider_name} - {model.model_name}")

    dbg.log_stage("model_selected", {
        "provider": model.provider_name,
        "model": model.model_name,
    })

    first_image = image_list[0] if image_list else None
    extract_method = "unknown"
    try:
        if first_image:
            if hasattr(model, 'extract') and first_image:
                try:
                    from pydantic import BaseModel
                    from typing import Any
                    import json as _json

                    class _RawOutput(BaseModel):
                        satirlar: list[Any] = []

                    extract_method = "extract_with_schema"
                    dbg.log_stage("extract_attempt", {"method": extract_method, "schema": "_RawOutput"})
                    response = model.extract(full_prompt, _RawOutput, image_data=first_image[0], mime_type=first_image[1])
                except Exception as extract_err:
                    extract_method = "raw_generate_fallback"
                    dbg.log_stage("extract_fallback", {
                        "original_error": str(extract_err),
                        "fallback_method": "raw_generate",
                    }, status="warning")
                    response = model.raw_generate(prompt=full_prompt)
            else:
                extract_method = "raw_generate"
                dbg.log_stage("extract_attempt", {"method": extract_method})
                response = model.raw_generate(prompt=full_prompt)
        else:
            extract_method = "raw_generate_text_only"
            dbg.log_stage("extract_attempt", {"method": extract_method})
            response = model.raw_generate(prompt=full_prompt)
    except Exception as e:
        dbg.log_error(e, context=f"extraction_{extract_method}")
        dbg.finish(success=False)
        raise

    dbg.log_stage("ai_response_received", {
        "method": extract_method,
        "has_structured_data": response.structured_data is not None,
        "content_length": len(response.content) if response.content else 0,
        "content_preview": (response.content[:1000] if response.content else ""),
        "structured_data": response.structured_data,
    })

    # Yapilandirilmis veriyi al
    raw_data = response.structured_data
    if not raw_data:
        raw_content = response.content if hasattr(response, 'content') else str(response)
        dbg.log_parsing("json_parse_attempt",
                       input_data={"raw_content_preview": raw_content[:2000],
                                   "raw_content_length": len(raw_content)})
        raw_data = _parse_json_response(raw_content)
        dbg.log_parsing("json_parse_result",
                       output_data=raw_data if raw_data else None,
                       error="parse_json_response bos dondu" if not raw_data else None)

    if not raw_data:
        dbg.log_stage("extraction_failed", {"reason": "raw_data bos"}, status="error")
        dbg.finish(success=False)
        st.error("Veri cikartma basarisiz oldu. Eski akis ile deniyor...")
        _process_create_legacy(combined_text, image_list, job_context)
        return

    with st.expander("Cikartilan Veri (Debug)"):
        st.json(raw_data)

    satirlar = raw_data.get("satirlar", [])

    dbg.log_stage("data_extracted", {
        "satirlar_count": len(satirlar),
        "raw_data_keys": list(raw_data.keys()) if isinstance(raw_data, dict) else str(type(raw_data)),
        "first_satir": satirlar[0] if satirlar else None,
    })

    st.markdown("### Onizleme")
    if satirlar:
        df = pd.DataFrame(satirlar)
        st.dataframe(df, use_container_width=True)
        st.write(f"**{len(satirlar)}** satir cikartildi")
    else:
        dbg.log_stage("no_satirlar", {"reason": "satirlar listesi bos"}, status="error")
        dbg.finish(success=False)
        st.warning("Veri cikartma sonucu bos. Eski akis deneniyor...")
        _process_create_legacy(combined_text, image_list, job_context)
        return

    st.divider()

    # Algoritmayı calistir
    with st.spinner("Is akisi (algoritma) calistiriliyor..."):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            dbg.log_stage("algorithm_start", {
                "func_id": func_id,
                "tmp_path": tmp_path,
                "raw_data_keys": list(raw_data.keys()) if isinstance(raw_data, dict) else "?",
            })
            run_algorithm(func_id, raw_data, tmp_path)
            with open(tmp_path, "rb") as f:
                excel_bytes = f.read()
            dbg.log_stage("algorithm_success", {"excel_size": len(excel_bytes)})
        except Exception as e:
            dbg.log_code_execution(
                code="run_algorithm(func_id, raw_data, tmp_path)",
                data=raw_data,
                error=str(e),
            )
            dbg.finish(success=False)
            st.error(f"Algoritma calistirma hatasi: {e}")
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
            # Fallback to legacy
            st.warning("Eski akis ile deneniyor...")
            _process_create_legacy(combined_text, image_list, job_context)
            return
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    dbg.finish(success=True, result={
        "excel_size": len(excel_bytes),
        "satirlar_count": len(satirlar),
    })
    _show_download_button(excel_bytes, "yeni")
    st.success("Excel hazir! (Algoritma tabanli)")


def _process_create_legacy(combined_text, image_list, job_context):
    """Eski akis: dogrudan AI ile Excel olusturma (3 adimli key dogrulama)."""
    prompt = _build_create_prompt(job_context, combined_text)
    first_image = image_list[0] if image_list else None
    if first_image:
        response = _call_ai_extract(prompt, VerifiedExcelData,
                                    image_data=first_image[0], mime_type=first_image[1])
    else:
        response = _call_ai_extract(prompt, VerifiedExcelData)

    data = response.structured_data

    ai_key_column = data.get("key_column", "")
    ai_key_values = data.get("key_values", [])

    with st.expander("AI Yaniti (Debug)"):
        st.json(data)

    st.markdown("### Onizleme")
    for sheet in data.get("sheetler", []):
        st.markdown(f"**{sheet['ad']}**")
        df = pd.DataFrame(sheet["veriler"], columns=sheet["basliklar"])
        st.dataframe(df, use_container_width=True)

    if ai_key_column and ai_key_values:
        missing_keys = _verify_keys(data, ai_key_column, ai_key_values)
        total_keys = len(ai_key_values)
        found_keys = total_keys - len(missing_keys)

        st.markdown("### Veri Butunlugu Kontrolu")
        st.write(f"Anahtar sutun: **{ai_key_column}**")
        st.write(f"Beklenen kayit: **{total_keys}** | Tabloda bulunan: **{found_keys}**")

        if missing_keys:
            st.error(f"AI {len(missing_keys)} kaydi tabloya ekleyemedi!")
            with st.expander(f"Eksik kayitlar ({len(missing_keys)} adet)", expanded=True):
                for mk in missing_keys:
                    st.write(f"  - `{mk}`")
        else:
            st.success("Tum kayitlar tabloda mevcut!")

    st.divider()
    excel_bytes = _create_excel_bytes(data)

    if ai_key_column and ai_key_values:
        missing_keys = _verify_keys(data, ai_key_column, ai_key_values)
        if missing_keys:
            excel_bytes = _add_failed_entries_to_excel(
                excel_bytes, ai_key_column, missing_keys
            )

    _show_download_button(excel_bytes, "yeni")
    st.success("Excel hazir!")


# ---------------------------------------------------------------------------
# Ana isleme fonksiyonu
# ---------------------------------------------------------------------------

def _process_all(mode, func, inputs, existing_info, reorganize_after):
    """Tum girisleri isleyen ana fonksiyon."""
    from core.debug_logger import AIDebugLogger

    if not _has_meaningful_input(inputs):
        st.error("Lutfen en az bir aractan veri girin!")
        return

    combined_text, image_list = _combine_inputs(inputs)

    if not combined_text.strip() and not image_list:
        st.error("Islenecek veri bulunamadi!")
        return

    job_context = _build_job_context(func)

    # Ust duzey akis logu
    dbg = AIDebugLogger("process_all", provider="tools_page", model="orchestrator")
    dbg.log_stage("process_start", {
        "mode": mode,
        "func_name": func.get("name", "Yok") if func else "Yok",
        "func_id": func.get("id", None) if func else None,
        "input_tools": list(inputs.keys()),
        "combined_text_length": len(combined_text),
        "combined_text_preview": combined_text[:500],
        "image_count": len(image_list),
        "has_existing_data": existing_info is not None,
        "reorganize_after": reorganize_after,
        "job_context_preview": job_context[:300],
    })

    try:
        with st.spinner("AI veriyi isliyor..."):

            # --- YENI OLUSTUR ---
            if mode == "create":
                func_id = func["id"] if func else None
                use_algorithm = False

                # Algoritma tabanlı akış kontrolü
                if func_id:
                    if algorithm_exists(func_id):
                        use_algorithm = True
                    else:
                        # Algoritma yok — otomatik oluşturmayı dene
                        router = st.session_state.get("router")
                        if router:
                            use_algorithm = _ensure_algorithm_ready(func_id, router)

                if use_algorithm and func_id:
                    # ===== YENİ AKIŞ: Algoritma tabanlı =====
                    _process_create_with_algorithm(
                        func, func_id, combined_text, image_list, job_context
                    )
                else:
                    # ===== ESKİ AKIŞ: Doğrudan AI ile (fallback) =====
                    _process_create_legacy(
                        combined_text, image_list, job_context
                    )

            # --- MEVCUT DOSYAYA EKLE ---
            elif mode == "append":
                prompt = _build_append_prompt(job_context, combined_text, existing_info)
                first_image = image_list[0] if image_list else None
                if first_image:
                    response = _call_ai_extract(prompt, ExcelData, image_data=first_image[0], mime_type=first_image[1])
                else:
                    response = _call_ai_extract(prompt, ExcelData)

                new_data = response.structured_data

                st.markdown("### Eklenecek Yeni Satirlar")
                for sheet in new_data.get("sheetler", []):
                    st.markdown(f"**{sheet['ad']}**")
                    df = pd.DataFrame(sheet["veriler"], columns=sheet["basliklar"])
                    st.dataframe(df, use_container_width=True)

                st.divider()

                # Algoritmik ekleme (AI degil)
                result_bytes = _append_to_excel(existing_info["raw_bytes"], new_data.get("sheetler", []))

                if reorganize_after:
                    result_bytes = _do_reorganize(job_context, result_bytes)

                _show_download_button(result_bytes, "eklenmis")
                st.success("Veriler mevcut dosyaya eklendi!")

            # --- MEVCUT DOSYADAN SIL ---
            elif mode == "delete":
                prompt = _build_delete_prompt(job_context, combined_text, existing_info)
                first_image = image_list[0] if image_list else None
                if first_image:
                    response = _call_ai_extract(prompt, DeleteInstruction, image_data=first_image[0], mime_type=first_image[1])
                else:
                    response = _call_ai_extract(prompt, DeleteInstruction)

                delete_data = response.structured_data

                key_col = delete_data.get("key_column", "")
                del_keys = delete_data.get("delete_keys", [])

                st.markdown("### Silinecek Satirlar")
                st.warning(f"Anahtar sutun: **{key_col}**")
                st.write(f"Silinecek degerler ({len(del_keys)} adet):")
                for k in del_keys:
                    st.write(f"  - `{k}`")

                st.divider()

                # Algoritmik silme
                result_bytes, deleted_count = _delete_from_excel(
                    existing_info["raw_bytes"], key_col, del_keys
                )

                if reorganize_after:
                    result_bytes = _do_reorganize(job_context, result_bytes)

                _show_download_button(result_bytes, "silindi")
                st.success(f"{deleted_count} satir silindi!")

            # --- MEVCUT DOSYAYI DUZENLE ---
            elif mode == "edit":
                prompt = _build_edit_prompt(job_context, combined_text, existing_info)
                first_image = image_list[0] if image_list else None
                if first_image:
                    response = _call_ai_extract(prompt, EditInstruction, image_data=first_image[0], mime_type=first_image[1])
                else:
                    response = _call_ai_extract(prompt, EditInstruction)

                edit_data = response.structured_data

                key_col = edit_data.get("key_column", "")
                edits = edit_data.get("edits", [])

                st.markdown("### Yapilacak Degisiklikler")
                st.info(f"Anahtar sutun: **{key_col}**")
                for e in edits:
                    st.write(f"  Satir `{e.get('key_value')}`: {json.dumps(e.get('changes', {}), ensure_ascii=False)}")

                st.divider()

                # Algoritmik duzenleme
                result_bytes, edited_count = _edit_excel(
                    existing_info["raw_bytes"], key_col, edits
                )

                if reorganize_after:
                    result_bytes = _do_reorganize(job_context, result_bytes)

                _show_download_button(result_bytes, "duzenlendi")
                st.success(f"{edited_count} hucre duzenlendi!")

    except Exception as e:
        dbg.log_error(e, context=f"process_all_{mode}")
        dbg.finish(success=False)
        st.error(f"Islem hatasi: {str(e)}")
        with st.expander("Detayli Hata"):
            st.code(traceback.format_exc())
        st.info("Detayli AI loglari icin soldaki menuden 'AI Debug Log' sayfasina gidin.")


def _do_reorganize(job_context, excel_bytes):
    """AI ile yeniden duzenleme. Key dogrulama ile veri butunlugu kontrolu."""
    st.info("AI ile yeniden duzenleme yapiliyor...")

    # Excel'i oku
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(excel_bytes), data_only=True)
    current_data = {"sheetler": []}
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        headers = []
        rows = []
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            row_list = [cell if cell is not None else "" for cell in row]
            if ri == 0:
                headers = [str(h) for h in row_list]
            else:
                rows.append(row_list)
        if headers:
            current_data["sheetler"].append({"ad": ws_name, "basliklar": headers, "veriler": rows})

    # Mevcut veriden key listesini cikar
    key_list = _extract_key_list_from_data(current_data)

    prompt = _build_reorganize_prompt(job_context, current_data, key_list)
    response = _call_ai_extract(prompt, VerifiedExcelData)
    reorganized = response.structured_data

    # Satir sayisi kontrolu
    orig_count = sum(len(s["veriler"]) for s in current_data.get("sheetler", []))
    new_count = sum(len(s["veriler"]) for s in reorganized.get("sheetler", []))

    if new_count != orig_count:
        st.warning(
            f"AI duzenleme sonrasi satir sayisi degisti ({orig_count} -> {new_count}). "
            f"Veri butunlugu riski! Orijinal dosya korunuyor."
        )
        return excel_bytes

    # Key dogrulama
    if key_list:
        missing = _verify_keys(reorganized, key_list["key_column"], key_list["key_values"])
        if missing:
            st.warning(
                f"Yeniden duzenleme sirasinda {len(missing)} kayit kayboldu. Orijinal korunuyor."
            )
            with st.expander("Kaybolan kayitlar"):
                for m in missing:
                    st.write(f"  - `{m}`")
            return excel_bytes

    result_bytes = _create_excel_bytes(reorganized)
    st.success("Yeniden duzenleme tamamlandi!")
    return result_bytes


# ---------------------------------------------------------------------------
# Sayfa giris noktasi
# ---------------------------------------------------------------------------

def show_tools_page():
    """Araclar sayfasi."""
    st.markdown('<div class="main-header">Araclar</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Veriyi Excel\'e donusturun</div>', unsafe_allow_html=True)

    # Router kontrolu
    if not st.session_state.router or not st.session_state.router.available_providers:
        st.error("Model router baslatilamadi. Lutfen API anahtarlarinizi ayarlayin.")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Ayarlara Git"):
                st.session_state.current_page = "settings"
                st.rerun()
        with col2:
            with st.expander("Gerekli Paketler"):
                st.code("pip install anthropic openai google-generativeai")
        st.stop()

    # --- 1. Is tanimi secimi ---
    active_business = st.session_state.active_business
    selected_func = None

    if active_business:
        functionalities = get_functionalities(active_business["id"])
        if functionalities:
            st.markdown("### Is Tanimi")
            func_options = {0: "Genel (Is tanimi olmadan)"}
            func_options.update({f["id"]: f["name"] for f in functionalities})
            selected_id = st.selectbox(
                "Hangi is icin calisiyorsunuz?",
                options=list(func_options.keys()),
                format_func=lambda x: func_options[x],
                key="tool_function_select",
            )
            if selected_id != 0:
                selected_func = next(f for f in functionalities if f["id"] == selected_id)
                st.info(f"**{selected_func['name']}** — {selected_func['description']}")

    st.divider()

    # --- 2. Islem modu ---
    st.markdown("### Islem Modu")
    mode = st.radio(
        "Ne yapmak istiyorsunuz?",
        options=["create", "append", "delete", "edit"],
        format_func=lambda x: MODE_LABELS[x],
        horizontal=True,
        key="operation_mode",
    )

    # --- 3. Mevcut Excel yukleme (olustur haricinde) ---
    existing_info = None
    if mode != "create":
        st.markdown("### Mevcut Excel Dosyasi")
        existing_file = st.file_uploader(
            "Islem yapilacak Excel dosyasini yukleyin",
            type=["xlsx", "xls"],
            key="existing_excel_file",
        )
        if existing_file:
            existing_info = _read_existing_excel(existing_file)
            if existing_info:
                with st.expander("Mevcut Veri Onizleme", expanded=False):
                    for sheet in existing_info.get("sheetler", []):
                        st.markdown(f"**{sheet['ad']}** ({len(sheet['veriler'])} satir)")
                        df = pd.DataFrame(sheet["veriler"], columns=sheet["basliklar"])
                        st.dataframe(df, use_container_width=True, height=200)
        else:
            st.warning("Lutfen islem yapilacak Excel dosyasini yukleyin.")

    st.divider()

    # --- 4. Arac secimi (coklu) ---
    st.markdown("### Giris Yontemi (birden fazla secilebilir)")

    selected_tool_ids = []
    tool_cols = st.columns(len(TOOLS))
    for idx, tool in enumerate(TOOLS):
        with tool_cols[idx]:
            if st.checkbox(
                f"{tool['name']}",
                key=f"tool_chk_{tool['id']}",
                help=tool["desc"],
            ):
                selected_tool_ids.append(tool["id"])

    if not selected_tool_ids:
        st.info("En az bir giris yontemi secin.")
        return

    st.divider()

    # --- 5. Giris alanlari ---
    inputs = {}
    if len(selected_tool_ids) == 1:
        inputs[selected_tool_ids[0]] = _show_tool_input(selected_tool_ids[0], selected_func, "single")
    else:
        tab_names = [f"{TOOL_LOOKUP[tid]['name']}" for tid in selected_tool_ids]
        tabs = st.tabs(tab_names)
        for tab, tid in zip(tabs, selected_tool_ids):
            with tab:
                inputs[tid] = _show_tool_input(tid, selected_func, f"tab_{tid}")

    st.divider()

    # --- 6. Opsiyonlar ve isle butonu ---
    reorganize_after = False
    if mode != "create":
        reorganize_after = st.checkbox(
            "Islem sonrasi AI ile bastan duzenle",
            help="Ekleme/silme/duzenleme sonrasi veri butunlugu bozulduysa AI ile yeniden organize edin. "
                 "AI mevcut verileri SILMEZ, yeni veri EKLEMEZ, sadece sirayi ve siniflandirmayi duzeltir.",
            key="reorganize_checkbox",
        )

    can_process = True
    if mode != "create" and existing_info is None:
        can_process = False

    if can_process:
        if st.button("Isle", type="primary", use_container_width=True, key="process_btn"):
            _process_all(mode, selected_func, inputs, existing_info, reorganize_after)
