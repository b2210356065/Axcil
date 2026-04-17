"""Kod uretme prompt sablonlari -- ozellikle openpyxl Excel kodu."""
from typing import Any, Optional
from pydantic import BaseModel


class CodeGenerationPromptBuilder:
    """
    Excel kodu uretme promptlari.

    Claude Sonnet icin optimize edilmis -- en iyi kod uretici.
    """

    @staticmethod
    def build_excel_generation(
        data_schema: dict[str, Any],
        business_context: dict[str, Any],
        sample_data: Optional[dict] = None,
        style_requirements: Optional[dict] = None,
    ) -> str:
        """
        openpyxl ile Excel olusturma kodu icin prompt.

        Args:
            data_schema: Veri yapisi semasi
            business_context: Is baglami
            sample_data: Ornek veri
            style_requirements: Stil gereksinimleri

        Returns:
            Kod uretme promptu
        """
        # Varsayilan stil gereksinimleri
        default_style = {
            "header_bg": "#1E3A5F",
            "header_fg": "#FFFFFF",
            "zebra_striping": True,
            "auto_width": True,
            "freeze_header": True,
            "number_format": "#,##0",
            "currency_format": '#,##0.00 "TL"',
            "date_format": "DD.MM.YYYY",
            "percentage_format": "0.00%",
        }
        style_reqs = {**default_style, **(style_requirements or {})}

        prompt = f"""Sen bir Python Excel otomasyon uzmanisin. openpyxl kutuphanesi ile
profesyonel, formatlI Excel dosyalari olusturan kod uretiyorsun.

<data_structure>
{data_schema}
</data_structure>

<sample_data>
{sample_data or "Veri yapisina uygun ornek veri olusturacaksin"}
</sample_data>

<business_context>
Is yeri: {business_context.get('business_name', 'Genel')}
Sektor: {business_context.get('sector', 'Genel')}
Kullanim amaci: {business_context.get('use_case', 'Veri kayit')}
</business_context>

<constraints>
ZORUNLU KURALLAR:

1. Fonksiyon imzasi: create_excel(data: dict, output_path: str) -> None
2. IZIN VERILEN importlar SADECE:
   - openpyxl (ve tum alt modulleri: styles, utils, worksheet, vb.)
   - datetime, date, timedelta
   - os.path (sadece path islemleri)
   - json, re, math, decimal
   - typing (type hint'ler icin)

3. KESINLIKLE YASAK -- bunlari ASLA KULLANMA:
   - pandas, numpy, xlsxwriter veya baska ucuncu parti kutuphane
   - eval(), exec(), compile()
   - subprocess, os.system(), os.popen()
   - __import__(), importlib
   - open() fonksiyonu (openpyxl wb.save() haric)
   - input(), raw_input()
   - globals(), locals() degistirme
   - socket, http, urllib (ag islemleri)
   ONEMLI: Tum veri isleme saf Python ile yapilmali (list, dict, for).
   pandas/numpy YOKTUR — kullanirsan kod calismaz!

4. openpyxl KURALLARI:
   - Satir/sutun numaralari 1-indexed (A1 = row=1, column=1)
   - Sayfa adi max 31 karakter, ozel karakter icermemeli
   - Buyuk veri setleri (>10000 satir) icin write_only=True kullan
   - Renk degerleri "RRGGBB" formati (# olmadan): "1E3A5F"
</constraints>

<style_requirements>
BASLIK SATIRI:
- Font: Calibri 11pt, kalin (bold=True)
- Arka plan: PatternFill(start_color="{style_reqs['header_bg'].lstrip('#')}", fill_type="solid")
- Yazi rengi: Font(color="{style_reqs['header_fg'].lstrip('#')}", bold=True)
- Hizalama: Alignment(horizontal="center", vertical="center", wrap_text=True)
- Kenarlik: Border(
    left=Side(style="medium", color="999999"),
    right=Side(style="medium", color="999999"),
    top=Side(style="medium", color="999999"),
    bottom=Side(style="medium", color="999999")
  )

VERI SATIRLARI:
{"- Zebra serit: Cift satirlar PatternFill(start_color='F2F2F2', fill_type='solid'), tek satirlar beyaz" if style_reqs['zebra_striping'] else "- Tek renk arka plan"}
- Kenarliklar: Border ile ince, gri (CCCCCC)
- Metin alanlari: sola hizali (horizontal="left")
- Sayisal alanlar: saga hizali (horizontal="right"), format: {style_reqs['number_format']}
- Para alanlari: saga hizali, format: {style_reqs['currency_format']}
- Tarih alanlari: ortali, format: {style_reqs['date_format']}
- Yuzde alanlari: ortali, format: {style_reqs['percentage_format']}
{"- Sutun genislikleri: Icerge gore otomatik ayarla (min: 10, max: 50 karakter)" if style_reqs['auto_width'] else ""}
{"- Baslik dondurusu: ws.freeze_panes = 'A2'" if style_reqs['freeze_header'] else ""}

TURKCE SAYI ve TARIH FORMATLAMA:
- Para birimi: Turkce Lira sembolunu kullan -> '#,##0.00 "TL"'
  (openpyxl number_format stringi olarak)
- Binlik ayirici: nokta (.) -- openpyxl'de '#,##0' otomatik yapar
- Ondalik ayirici: virgul (,) -- openpyxl'de locale bagimli, format stringi yeterli
- Tarih: DD.MM.YYYY (Turkiye standardi) -> number_format = "DD.MM.YYYY"
- Saat: HH:MM (24 saat formati)

OZET/TOPLAM SATIRI:
- Son satirda toplam/ozet satiri ekle (eger sayisal sutunlar varsa)
- Toplam satiri stil: kalin font, ust cizgi (top border medium)
- SUM formullerini kullan: =SUM(C2:C{{son_satir}})
- Toplam satir arka plan: acik mavi PatternFill(start_color="D6EAF8", fill_type="solid")

YAZDIRMA AYARLARI:
- ws.sheet_properties.pageSetUpPr.fitToPage = True
- ws.page_setup.orientation = "landscape"
- ws.page_setup.fitToWidth = 1
- ws.page_setup.fitToHeight = 0 (sayfa sayisi siniri yok)
- ws.page_setup.paperSize = ws.PAPERSIZE_A4
</style_requirements>

<code_structure>
Kod su yapida olmali:

```python
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os
import math

def create_excel(data: dict, output_path: str) -> None:
    \"\"\"
    [Kisa aciklama]

    Args:
        data: Veri dictionary'si
        output_path: Cikti dosya yolu
    \"\"\"
    # 1. Workbook olustur
    wb = Workbook()
    ws = wb.active
    ws.title = "..."  # max 31 karakter

    # 2. Stil tanimlari (bir kez tanimla, tekrar kullan)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # ... diger stiller

    # 3. Baslik satiri
    headers = [...]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = ...

    # 4. Veri satirlari
    for row_idx, item in enumerate(data_list, start=2):
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Tip bazli formatlama
            # Zebra sirit

    # 5. Otomatik sutun genisligi
    for col_idx in range(1, len(headers) + 1):
        max_length = max(len(str(cell.value or ""))
                        for cell in ws[get_column_letter(col_idx)])
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # 6. Toplam satiri (eger sayisal alan varsa)
    # ...

    # 7. Baslik dondur + yazdirma ayarlari
    ws.freeze_panes = "A2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    # 8. Kaydet
    wb.save(output_path)
```
</code_structure>

<output_format>
- SADECE Python kodu uret.
- Ürettiğin kod test senaryolarını İÇERMEMELİ genel iş mantığı için kullanılabilmeli.
- Aciklama, markdown, yorum (docstring haric) ekleme.
- Kod calisir durumda olmali.
- Test edilebilir olmali.
- Fonksiyon disinda kod yazma (import'lar haric).
</output_format>

create_excel fonksiyonunu yukaridaki gereksinimlere gore uret:
"""
        return prompt

    @staticmethod
    def build_formula_generation(
        formula_description: str,
        data_context: dict[str, Any],
    ) -> str:
        """Excel formulu uretme promptu."""
        prompt = f"""Sen bir Excel formul uzmanisin. Aciklamaya gore Excel formulu uretiyorsun.

<formula_requirement>
{formula_description}
</formula_requirement>

<data_context>
Sutunlar: {data_context.get('columns', [])}
Veri araligi: {data_context.get('data_range', 'A2:Z100')}
Ozel kosullar: {data_context.get('conditions', 'Yok')}
</data_context>

<formula_rules>
1. Excel formul sozdizimini kullan (=SUM, =IF, =VLOOKUP, vb.)
2. Turkce karakter ICERME (formul icinde Turkce fonksiyon adi kullanma)
3. Hucre referanslari A1 formatinda
4. Karmasik formuller icin aciklama ekle
5. Sadece formulu dondur (= ile basla)
6. Para hesaplamalarinda: ROUND fonksiyonu ile 2 ondalik basamaga yuvarla
7. Tarih hesaplamalarinda: DATE, YEAR, MONTH, DAY fonksiyonlarini kullan
</formula_rules>

Formulu uret:
"""
        return prompt

    @staticmethod
    def build_data_transformation(
        source_format: str,
        target_format: str,
        transformation_rules: list[str],
    ) -> str:
        """Veri donusturme kodu promptu."""
        prompt = f"""Veri donusturme fonksiyonu uret.

<source_format>
{source_format}
</source_format>

<target_format>
{target_format}
</target_format>

<transformation_rules>
{chr(10).join(f"{i+1}. {rule}" for i, rule in enumerate(transformation_rules))}
</transformation_rules>

<requirements>
- Fonksiyon: transform_data(source: dict) -> dict
- Hata yonetimi: try-except ile guvenli donusum
- Tip donusumleri: str -> int, str -> date, vb.
- Turkce sayi formatlari: "1.234,56" -> 1234.56
- Turkce tarih formatlari: "15.03.2025" -> date(2025, 3, 15)
- Validation: Donusturulen veri hedef formata uygun mu kontrol et
- YASAK: eval, exec, subprocess, __import__
</requirements>

Python fonksiyonunu uret:
"""
        return prompt

    @staticmethod
    def build_with_examples(
        task_description: str,
        examples: list[dict],
        language: str = "python",
    ) -> str:
        """Orneklerle kod uretme (Few-Shot)."""
        examples_section = "\n<examples>\n"

        for i, example in enumerate(examples, 1):
            examples_section += f"\nOrnek {i}:\n"
            examples_section += f"Gorev: {example.get('task', '')}\n"
            examples_section += f"```{language}\n{example.get('code', '')}\n```\n"

        examples_section += "</examples>\n"

        prompt = f"""{examples_section}

Simdi asagidaki gorevi yukaridaki orneklere benzer sekilde coz:

<task>
{task_description}
</task>

<safety_constraints>
KESINLIKLE YASAK:
- eval(), exec(), compile()
- subprocess, os.system(), os.popen()
- __import__(), importlib
- open() (openpyxl wb.save haric)
- Ag islemleri (socket, http, urllib)
</safety_constraints>

{language} kodunu uret:
"""
        return prompt


# ---------------------------------------------------------------------------
# Zenginlestirme Prompt Builder
# ---------------------------------------------------------------------------

class EnrichmentPromptBuilder:
    """Is tanimi zenginlestirme promptlari (prompt_engineering_strategies.md Bolum 2-3)."""

    @staticmethod
    def build_enrichment(name: str, description: str, sector: str,
                         data_types: list[str]) -> str:
        """Ilk zenginlestirme promptu (Bolum 2.3)."""
        return f"""Sen bir iş süreci analisti ve bilgi mimarısın.

<görev>
Kullanıcının verdiği kısa iş tanımını analiz et. Bu tanımdan kapsamlı, yapılandırılmış
ve hem insan hem yapay zeka tarafından net anlaşılacak bir iş tanımı belgesi oluştur.
</görev>

<kullanıcı_girdisi>
İş adı: {name}
Açıklama: {description}
Sektör: {sector}
Kabul edilen veri tipleri: {', '.join(data_types)}
</kullanıcı_girdisi>

<zenginleştirme_talimatları>

================================================================
ÖNCELİKLE: İŞ TANIMINI ANLA
================================================================

Kullanıcının verdiği iş tanımını kullanıcının gözünden düşün. Kullanıcı bu işi günlük
hayatında nasıl yapıyor? Hangi verilerle çalışıyor? Hangi zorluklar yaşıyor? Kullanıcının
bu tanımla gerçekte ne demek istediğini, hangi iş akışını dijitalleştirmek istediğini,
bu işin sonunda nasıl bir Excel beklediğini kendi kelimeleriyle — teknik dil kullanmadan,
bir iş insanının anlayacağı dilde — en az bir paragraf olarak yaz. Bu paragraf senin
iş tanımını ne kadar iyi anladığının kanıtı olacak. İş tanımını yetersiz düşünürsen
yanlış çıkarımlar yapacaksın, o yüzden bu adımı atlamadan ve kısaltmadan tamamla.

================================================================
BÖLÜM 1 — İŞ ÖZETİ
================================================================

Yukarıdaki anlamanın üzerine inşa ederek iş tanımını kullanıcının perspektifinden
akıcı ve tutarlı bir şekilde özetle. Bu özet:
- Kimin için, ne amaçla, hangi süreçte kullanıldığını açıklamalı
- Kullanıcının günlük iş akışındaki yerini tanımlamalı
- Hem insan hem AI tarafından okunabilir bir dilde yazılmalı
- İş tanımının kapsamını net olarak çizmeli

================================================================
BÖLÜM 2 — VERİ YAPISI (SÜTUNLAR)
================================================================

Her sütun için:
- Sütun adı (Türkçe, anlaşılır)
- Veri tipi (metin/sayı/tarih/para/yüzde/seçim/formül)
- Zorunlu mu? (evet/hayır)
- Varsayılan değer (varsa)
- Formül (varsa — Excel formül mantığı olarak yaz)
- Doğrulama kuralı (varsa)
- Açıklama (kullanıcının bu sütunu nasıl anlaması gerektiği)

================================================================
BÖLÜM 3 — İŞ KURALLARI
================================================================

- Matematiksel tutarlılık kuralları (örn: Toplam = Ara Toplam + KDV)
- Format kuralları (tarih formatı, para birimi, ondalık basamak)
- Mantıksal kısıtlamalar (örn: tarih gelecekte olamaz)
- Veri tipleri arası ilişkiler

================================================================
BÖLÜM 4 — GÖRSEL VE SUNUM
================================================================

- Sayfa adı önerisi
- Sıralama tercihi (hangi sütuna göre, ASC/DESC)
- Gruplama (varsa)
- Toplam/özet satırı gereken sütunlar
- Özel formatlama notları

================================================================
BÖLÜM 5 — ÖRNEK SENARYOLAR
================================================================

Yukarıdaki BÖLÜM 2, 3 ve 4'teki maddelerden referans alarak somut senaryolar yaz.
Her senaryo iş tanımının içerisindeki belirli bir maddeyi veya madde grubunu somutlaştırır.

Her senaryo şu formatta olsun:

  Senaryo: [Kısa başlık]
  Referans: [Bölüm X, madde Y — hangi kural/sütun/özellik test ediliyor]
  Girdi: [Kullanıcının ne verdiğini tarif et — foto/metin/PDF olarak]
  Beklenen çıktı: [Satırların Excel'de nasıl görüneceği — tablo satırı olarak]
  Açıklama: [Bu senaryonun neden önemli olduğu, hangi iş kuralını gösterdiği]

Senaryolar iş tanımını somutlaştırmalı — soyut kalmamalı.
Toplam 3-7 senaryo yeterli — abartma.
Senaryolar belgenin EN ALTINDA yer almalı.

</zenginleştirme_talimatları>

<kalite_kuralları>
- Tereddüt ifadeleri kullanma ("belki", "muhtemelen", "sanırım").
- Sütun isimlerini Türkçe ve anlaşılır tut.
- Formülleri Excel formül mantığı ile yaz (Python değil).
- Her kuralın gerekçesini kullanıcının anlayacağı dilde aç.
- İş tanımı hem insanın okuyup "evet bu benim işim" diyeceği hem de
  AI'ın şema/kural/mantık çıkaracağı bir dilde ve yapıda sunulmalı.
</kalite_kuralları>

<çıktı_formatı>
Sonucu SADECE aşağıdaki JSON formatında döndür — ek yorum ekleme:
{{
  "iş_özeti": "...",
  "sütunlar": [
    {{"ad": "...", "tip": "...", "format": "...", "zorunlu": true, "varsayılan": null, "formül": null, "doğrulama": "...", "açıklama": "..."}}
  ],
  "iş_kuralları": ["...", "..."],
  "sunum": {{"sayfa_adı": "...", "sıralama": {{"sütun": "...", "yön": "ASC"}}, "gruplama": null, "toplam_sütunları": ["..."]}},
  "senaryolar": [
    {{"başlık": "...", "referans": "...", "girdi": "...", "beklenen_çıktı": {{}}, "açıklama": "..."}}
  ]
}}
</çıktı_formatı>
"""

    @staticmethod
    def build_iterative_enrichment(name: str, description: str, sector: str,
                                   data_types: list[str],
                                   failed_attempts: list[dict]) -> str:
        """Iteratif zenginlestirme promptu (Bolum 3.3)."""
        attempts_text = ""
        for attempt in failed_attempts:
            attempts_text += f"""
--- Deneme {attempt.get('attempt_number', '?')} ({attempt.get('created_at', '?')}): ---

Üretilen tanım:
{attempt.get('enriched_definition', '{}')}

Kullanıcının eksikler ve önerileri:
\"{attempt.get('user_feedback', 'Belirtilmemiş')}\"

--- / ---
"""

        return f"""Sen bir iş süreci analisti ve bilgi mimarısın.

<görev>
Kullanıcı daha önce üretilen iş tanımlarından memnun kalmadı.
Kullanıcının orijinal isteğini, başarısız denemeleri ve geri bildirimleri
inceleyerek DAHA İYİ bir zenginleştirilmiş iş tanımı oluştur.
</görev>

================================================================
ADIM 1: KULLANICININ ORİJİNAL İSTEĞİNİ ANLA
================================================================

<orijinal_istek>
İş adı: {name}
Açıklama: {description}
Sektör: {sector}
Veri tipleri: {', '.join(data_types)}
</orijinal_istek>

Öncelikle kullanıcının orijinal isteğini dikkatlice oku. Kullanıcının bu iş
tanımıyla GERÇEKTE ne yapmak istediğini anlamaya çalış. Kullanıcının bakış
açısından düşün — teknik değil, iş süreci odaklı. Kullanıcının günlük hayatında
bu işi nasıl yaptığını, hangi sorunlarla karşılaştığını, bu sistemden ne
beklediğini kendi kelimeleriyle en az bir paragraf olarak yaz.

================================================================
ADIM 2: BAŞARISIZ DENEMELERİ İNCELE
================================================================

<başarısız_denemeler>
Aşağıdaki denemeler kullanıcı tarafından reddedildi.
Her denemeyi kullanıcının geri bildirimiyle birlikte oku.
Kullanıcının ne demek istediğini DÜŞÜN — sadece kelimelerine değil,
arkasındaki niyete odaklan.

{attempts_text}
</başarısız_denemeler>

Şimdi başarısız denemeleri ve kullanıcının geri bildirimlerini düşün:
- Her denemenin neden başarısız olduğunu belirle.
- Kullanıcının geri bildirimlerindeki ortak temaları çıkar.
- Önceki denemelerde tekrarlanan hataları listele — bunları ASLA tekrarlama.

================================================================
ADIM 3: YENİ İŞ TANIMI OLUŞTUR
================================================================

KURALLAR:
- Kullanıcının geri bildirimlerindeki HER maddeyi adresle
- Daha önce reddedilen yaklaşımları TEKRARLAMA
- Kullanıcının öneride bulundukları kısımları ÖNCELİKLE uygula

Yeni tanımı oluştur — BÖLÜM 1 (İş Özeti), BÖLÜM 2 (Sütunlar), BÖLÜM 3 (İş Kuralları),
BÖLÜM 4 (Sunum), BÖLÜM 5 (Senaryolar) yapısında.

================================================================
ADIM 4: KALİTE KONTROL
================================================================

Yeni tanımı döndürmeden önce:
- Kullanıcının her geri bildirim maddesini adresliyor musun?
- Önceki denemelerin reddedilme sebeplerinden herhangi birini tekrarlıyor musun?
- Senaryolar iş tanımının içerisindeki maddelere referans veriyor mu?

<çıktı_formatı>
Sonucu SADECE aşağıdaki JSON formatında döndür — ek yorum ekleme:
{{
  "iş_özeti": "...",
  "sütunlar": [
    {{"ad": "...", "tip": "...", "format": "...", "zorunlu": true, "varsayılan": null, "formül": null, "doğrulama": "...", "açıklama": "..."}}
  ],
  "iş_kuralları": ["...", "..."],
  "sunum": {{"sayfa_adı": "...", "sıralama": {{"sütun": "...", "yön": "ASC"}}, "gruplama": null, "toplam_sütunları": ["..."]}},
  "senaryolar": [
    {{"başlık": "...", "referans": "...", "girdi": "...", "beklenen_çıktı": {{}}, "açıklama": "..."}}
  ]
}}
</çıktı_formatı>
"""


# ---------------------------------------------------------------------------
# Algoritma Uretim Prompt Builder
# ---------------------------------------------------------------------------

class AlgorithmPromptBuilder:
    """Is akisi (algoritma) uretim promptlari (prompt_engineering_strategies.md Bolum 4-5)."""

    @staticmethod
    def build_algorithm_generation(enriched_definition: dict) -> str:
        """Algoritma uretim promptu — 6 asamali Reflexion (Bolum 4.3)."""
        import json as _json

        enriched_json = _json.dumps(enriched_definition, ensure_ascii=False, indent=2)

        # Sutun listesi
        columns = enriched_definition.get("sütunlar",
                   enriched_definition.get("sutunlar", []))
        col_names = [c.get("ad", "") for c in columns]
        col_list_str = ", ".join(col_names)

        return f"""Sen bir Python Excel otomasyon uzmanısın. openpyxl kütüphanesiyle çalışan,
güvenli ve profesyonel Excel dosyaları oluşturan kod yazıyorsun.

<görev>
Aşağıdaki iş tanımına göre create_excel fonksiyonu üret.
Bu fonksiyon bir kez yazılacak ve aynı iş tanımı için tekrar tekrar kullanılacak.
Fonksiyon SADECE veriyi Excel'e yazmaktan sorumlu — veri işlemez, dönüştürmez.
</görev>

<iş_tanımı>
{enriched_json}
</iş_tanımı>

================================================================
AŞAMA 1: İŞ TANIMINI ANLA
================================================================

İş tanımını ve içerisindeki örnek senaryoları dikkatlice oku.
Bu iş tanımının ne yaptığını, hangi verileri işleyeceğini, hangi formatlama
ve hesaplama kurallarına uyacağını TAM OLARAK anlamaya çalış.

<anlama>
[İş tanımını kendi kelimleriyle en az bir paragraf olarak açıkla.
Bu işin amacını, veri akışını, kritik kuralları ve özel durumlarını
KENDİ ANLADIĞIN ŞEKİLDE yaz. Bu paragraf senin iş tanımını ne kadar
iyi kavradığının kanıtı olacak — yetersiz anlarsan yanlış kod yazarsın.]

Teknik detaylar:
- Kaç sütun var ve tipleri neler? (liste)
- Formül içeren sütunlar hangileri? (liste)
- Toplam/özet satırı gerekli mi? (evet/hayır + hangi sütunlar)
- Özel formatlama gereken alanlar? (liste)
</anlama>

================================================================
AŞAMA 2: GELİŞMİŞ SENARYOLAR TANIMLA
================================================================

İş tanımındaki her önemli özellik/madde için 3-5 arası gelişmiş test
senaryosu tanımla. Her senaryo için:
1. Senaryoyu tanımla
2. Bu senaryonun iş tanımına göre NASIL çözülmesi gerektiğini düşün
3. Çözümün arkasındaki mantığı not et

<senaryolar>
Senaryo 1: [Başlık]
  Durum: [Hangi kural test ediliyor]
  Girdi: data = {{"satirlar": [{{...}}]}}
  Beklenen: [Hangi hücrede ne değer/formül/stil]
  Çözüm mantığı: [Kodda ne olmalı]

[...3-5 senaryo...]
</senaryolar>

================================================================
AŞAMA 3: KODU YAZ
================================================================

<kısıtlamalar>
FONKSİYON İMZASI: create_excel(data: dict, output_path: str) -> None

GİRDİ FORMATI: data["satirlar"] bir liste — her eleman bir dict.
Sütun isimleri: {col_list_str}

İZİN VERİLEN İMPORTLAR (SADECE bunlar):
- openpyxl (tüm alt modülleri: styles, utils, worksheet, chart)
- datetime, date, timedelta
- os.path (SADECE path işlemleri)
- json, re, math, decimal

KESİNLİKLE YASAK — bunları ASLA kullanma:
- pandas, numpy, xlsxwriter veya başka üçüncü parti kütüphane
- eval(), exec(), compile()
- subprocess, os.system(), os.popen()
- __import__(), importlib
- open() (openpyxl wb.save() HARİÇ)
- input(), raw_input()
- globals(), locals() değiştirme
- socket, http, urllib

ÖNEMLİ: Tüm veri işleme saf Python ile yapılmalı (list, dict, for döngüsü).
pandas, numpy gibi kütüphaneler YOKTUR — kullanırsan kod çalışmaz!
</kısıtlamalar>

<stil_gereksinimleri>
BAŞLIK SATIRI:
- Font: Calibri 11pt, bold, beyaz (FFFFFF)
- Arka plan: Koyu mavi (1E3A5F), solid fill
- Hizalama: Ortalı, wrap_text=True
- Kenarlık: Medium, gri (999999)

VERİ SATIRLARI:
- Zebra: Çift satırlar açık gri (F2F2F2), tek satırlar beyaz
- Kenarlık: İnce, gri (CCCCCC)
- Metin: sola hizalı
- Sayı: sağa hizalı, format #,##0
- Para: sağa hizalı, format #,##0.00 "TL"
- Tarih: ortalı, format DD.MM.YYYY
- Yüzde: ortalı, format 0.00%

TOPLAM SATIRI (varsa):
- Font: Bold
- Üst çizgi: medium border
- Arka plan: Açık mavi (D6EAF8)
- SUM formülleri: =SUM(C2:C{{son_satır}})

SAYFA AYARLARI:
- freeze_panes = "A2"
- Landscape, fit to page, A4
</stil_gereksinimleri>

<kod>
[Fonksiyonu buraya yaz]
</kod>

================================================================
AŞAMA 4: KARMAŞIK SENARYOLARLA ZİHİNSEL TEST
================================================================

Fonksiyonun AŞAMA 2'de hazırladığın senaryoları doğru işleyip
işlemediğini kontrol et. Her senaryoyu kodun üzerinden zihinsel
olarak adım adım geçir:

<karmaşık_test_sonuçları>
Senaryo 1: GEÇTİ / KALDI
Senaryo 2: GEÇTİ / KALDI
[...tüm senaryolar...]
</karmaşık_test_sonuçları>

Eğer herhangi bir senaryo KALDIYSA, sorunu belirle, kodu düzelt
ve testleri tekrarla. (Maks 3 deneme)

================================================================
AŞAMA 5: BASİT SENARYOLARLA EK DOĞRULAMA
================================================================

<basit_testler>
1. Boş satırlar listesi → Sadece başlık satırı olmalı ✓/✗
2. Tek satırlı veri → 1 başlık + 1 veri satırı ✓/✗
3. 5 satırlı standart veri → Tüm stiller doğru ✓/✗
4. null değer içeren satır → Hata vermemeli ✓/✗
5. Çok uzun metin içeren alan → wrap_text çalışmalı ✓/✗
6. Negatif sayı → Format doğru uygulanmalı ✓/✗
7. Tarih alanı boş → Hata vermemeli ✓/✗
8. Formül sütunu → Excel formülü yazılmalı ✓/✗
9. 100 satırlık veri → Otomatik genişlik çalışmalı ✓/✗
10. Türkçe karakter içeren sütun adı → Doğru yazılmalı ✓/✗
Sonuç: {{geçen}}/{{toplam}}
</basit_testler>

================================================================
AŞAMA 6: SONUÇ
================================================================

Fonksiyon hem karmaşık hem basit testlerin TAMAMINI geçiyorsa,
aşağıdaki JSON formatında döndür:

{{"status": "success", "code": "[create_excel fonksiyonunun TAM kodu — importlar dahil]", "test_summary": {{"senaryo_testleri": {{"gecen": 0, "toplam": 0}}, "ek_testler": {{"gecen": 0, "toplam": 10}}}}, "notlar": "..."}}

Eğer 3 denemede de testler geçemiyorsa:

{{"status": "failure", "son_kod": "[En son denenen kodun TAMI]", "basarisiz_testler": [{{"senaryo": "...", "beklenen": "...", "gerceklesen": "...", "tahmin_edilen_sebep": "..."}}], "oneri": "...", "deneme_sayisi": 3}}

SADECE JSON DÖNDÜR — başka bir şey yazma!
"""

    @staticmethod
    def build_algorithm_iteration(enriched_definition: dict,
                                  failed_attempts: list[dict]) -> str:
        """Basarisiz algoritmadan sonra yeniden uretim promptu (Bolum 5.3)."""
        import json as _json

        enriched_json = _json.dumps(enriched_definition, ensure_ascii=False, indent=2)

        attempts_text = ""
        for attempt in failed_attempts:
            attempts_text += f"""
--- Deneme {attempt.get('attempt_number', '?')} ({attempt.get('created_at', '?')}): ---
Kod:
```python
{attempt.get('code', '')}
```

Başarısız testler:
{attempt.get('test_results', 'Belirtilmemiş')}

AI başarısızlık raporu:
{attempt.get('ai_failure_report', 'Yok')}

Kullanıcının eksikler ve önerileri:
\"{attempt.get('user_feedback', 'Yok')}\"
--- / ---
"""

        return f"""Sen bir Python Excel otomasyon uzmanısın.

<görev>
Daha önce bu iş tanımı için algoritma üretildi ancak başarısız oldu.
Geçmiş denemeleri ve hata raporlarını inceleyerek DOĞRU çalışan bir
create_excel fonksiyonu üret.
</görev>

<iş_tanımı>
{enriched_json}
</iş_tanımı>

<hata_geçmişi>
AŞAĞIDAKİ HATALARI KESİNLİKLE TEKRARLAMA.
{attempts_text}
</hata_geçmişi>

Şimdi yeni bir create_excel fonksiyonu üret. Bölüm 4.3'teki AŞAMA 1-6
prosedürünü AYNEN uygula, ancak şu EK kurallara da uy:

- AŞAMA 1'de iş tanımını anlarken geçmiş hataları da göz önünde bulundur
- AŞAMA 2'de senaryo tanımlarken ÖNCEKİ denemelerde başarısız olan
  senaryoları ÖNCELİKLE dahil et
- AŞAMA 3'te kod yazarken geçmiş denemelerdeki hatalı yaklaşımları KULLANMA
- Kullanıcının eksikler ve önerilerindeki HER maddeyi adresle

SADECE JSON formatında döndür:
Başarılıysa: {{"status": "success", "code": "...", "test_summary": {{}}, "notlar": "..."}}
Başarısızsa: {{"status": "failure", "son_kod": "...", "basarisiz_testler": [...], "oneri": "...", "deneme_sayisi": 3}}
"""


# ---------------------------------------------------------------------------
# Calisma Zamani Cikarma Prompt Builder
# ---------------------------------------------------------------------------

class RuntimeExtractionPromptBuilder:
    """Calisma zamaninda veri cikarma promptlari (Bolum 6)."""

    @staticmethod
    def build_runtime_extraction(enriched_definition: dict,
                                 input_type: str = "metin") -> str:
        """Zenginlestirilmis tanimdan sutun bilgisi alarak JSON cikarma promptu."""
        import json as _json

        columns = enriched_definition.get("sütunlar",
                   enriched_definition.get("sutunlar", []))

        col_info = []
        col_names = []
        for c in columns:
            name = c.get("ad", "")
            tip = c.get("tip", "text")
            formul = c.get("formül", c.get("formul"))
            col_names.append(name)
            if formul:
                col_info.append(f"- {name} ({tip}) — FORMÜL: Bu sütunu null bırak, Excel hesaplayacak")
            else:
                col_info.append(f"- {name} ({tip})")

        col_info_text = "\n".join(col_info)
        col_names_text = ", ".join(col_names)

        example_row = {}
        for c in columns:
            name = c.get("ad", "")
            tip = c.get("tip", "text")
            formul = c.get("formül", c.get("formul"))
            if formul:
                example_row[name] = None
            elif tip == "date":
                example_row[name] = "YYYY-MM-DD"
            elif tip in ("number", "currency", "percentage"):
                example_row[name] = 0
            else:
                example_row[name] = "..."

        example_json = _json.dumps({"satirlar": [example_row]}, ensure_ascii=False, indent=2)

        return f"""Aşağıdaki {input_type} içeriğinden verileri çıkar.

<sütunlar>
{col_info_text}
</sütunlar>

<çıktı_formatı>
{example_json}
</çıktı_formatı>

<kurallar>
- TÜM satırları çıkar, HİÇBİRİNİ atlama
- Sütun isimlerini AYNEN kullan: {col_names_text}
- Sayıları sayı olarak döndür (string değil)
- Tarihleri YYYY-MM-DD formatında döndür
- Boş/belirsiz alanlar: null
- Para tutarlarını ondalık sayı olarak yaz (sembol ekleme)
- Formül sütunlarını null bırak (Excel hesaplayacak)
- SADECE JSON döndür, ek açıklama ekleme
- Tereddüt ifadeleri kullanma ("aslında", "bekle", "tekrar düşüneyim")
</kurallar>

<doğrulama>
Çıktını döndürmeden önce:
1. Tüm satırları saydın mı?
2. Tüm sütunlar her satırda mevcut mu?
3. Veri tipleri doğru mu (sayı=sayı, tarih=tarih)?
Eksik varsa düzelt.
</doğrulama>
"""


# ---------------------------------------------------------------------------
# Yardimci fonksiyonlar
# ---------------------------------------------------------------------------

def quick_excel_code_prompt(
    data_schema: dict,
    context: dict = None,
) -> str:
    """Hizli Excel kodu promptu."""
    return CodeGenerationPromptBuilder.build_excel_generation(
        data_schema=data_schema,
        business_context=context or {},
    )
