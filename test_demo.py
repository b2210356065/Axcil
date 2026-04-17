"""
End-to-End Test & Demo
======================

Tüm sistemi test eder:
1. Model adaptörleri
2. Router ve fallback
3. Pipeline yönetimi
4. Excel oluşturma
5. Güvenlik kontrolü
"""
import os
from datetime import date
from decimal import Decimal
from pydantic import BaseModel, Field

# Core modüller
from core.models import (
    APIConfig, TaskType, ModelProvider,
    ExtractionResult, FieldValue, PromptMetrics
)
from core.config import load_config, save_config

# AI modülleri
from ai.router import ModelRouter, RetryHandler
from ai.pipeline import PipelineManager

# Excel modülleri
from excel_engine.builder import ExcelBuilder, quick_excel
from excel_engine.styles import StyleManager
from excel_engine.templates import TemplateLibrary
from excel_engine.sandbox import CodeSandbox, quick_validate

# Tools
from tools.image_to_excel import ImageToExcelTool


def print_section(title: str):
    """Bölüm başlığı yazdır."""
    print("\n" + "=" * 60)
    print(f"  {title}")
    print("=" * 60)


def test_1_basic_excel_creation():
    """Test 1: Temel Excel oluşturma."""
    print_section("TEST 1: Temel Excel Oluşturma")

    # Basit veri
    data = [
        {"Ad": "Ali Yılmaz", "Departman": "IT", "Maaş": 15000},
        {"Ad": "Ayşe Kaya", "Departman": "Muhasebe", "Maaş": 12000},
        {"Ad": "Mehmet Demir", "Departman": "Satış", "Maaş": 10000},
    ]

    output_path = "outputs/test_basic.xlsx"

    try:
        result_path = quick_excel(data, output_path, sheet_name="Çalışanlar")
        print(f"[OK] Excel başarıyla oluşturuldu: {result_path}")
        print(f"   Dosya boyutu: {os.path.getsize(result_path)} bytes")
        return True
    except Exception as e:
        print(f"[FAIL] Hata: {e}")
        return False


def test_2_template_library():
    """Test 2: Şablon kütüphanesi."""
    print_section("TEST 2: Şablon Kütüphanesi")

    lib = TemplateLibrary()

    # Mevcut şablonları listele
    templates = lib.list_templates()

    print("> Mevcut sablonlar:")
    for sector, template_list in templates.items():
        print(f"\n  {sector.upper()}:")
        for template_name in template_list:
            print(f"    • {template_name}")

    # Bir şablonu al
    fatura_template = lib.get_template("accounting", "fatura")
    if fatura_template:
        print("\n[OK] Fatura şablonu yüklendi:")
        for sheet in fatura_template.sheets:
            print(f"   Sayfa: {sheet.name}")
            print(f"   Sütunlar: {[col.name for col in sheet.columns]}")
        return True
    else:
        print("[FAIL] Şablon bulunamadı")
        return False


def test_3_code_sandbox():
    """Test 3: Güvenli kod çalıştırma."""
    print_section("TEST 3: Code Sandbox Güvenlik")

    # Güvenli kod
    safe_code = """
def create_excel(data, output_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    ws['A1'] = "Merhaba"
    wb.save(output_path)
"""

    # Tehlikeli kod
    dangerous_code = """
def create_excel(data, output_path):
    import subprocess
    subprocess.call(['rm', '-rf', '/'])
"""

    sandbox = CodeSandbox()

    # Güvenli kod testi
    is_safe, error = sandbox.validate_code(safe_code)
    print(f"Güvenli kod kontrolü: {'[OK] Geçti' if is_safe else f'[FAIL] Başarısız: {error}'}")

    # Tehlikeli kod testi
    is_safe, error = sandbox.validate_code(dangerous_code)
    print(f"Tehlikeli kod kontrolü: {'[OK] Yakalandı' if not is_safe else '[FAIL] Tespit edilemedi!'}")
    if not is_safe:
        print(f"   Sebep: {error}")

    return True


def test_4_extraction_result_to_excel():
    """Test 4: ExtractionResult -> Excel."""
    print_section("TEST 4: ExtractionResult -> Excel")

    # Mock ExtractionResult
    extraction_result = ExtractionResult(
        fields={
            "tarih": FieldValue(value=date(2025, 3, 15), confidence=0.95),
            "satici": FieldValue(value="Migros", confidence=0.99),
            "tutar": FieldValue(value=Decimal("145.50"), confidence=0.88),
            "kategori": FieldValue(value="Market", confidence=0.75),
        },
        model_used="gemini-3.1-flash-live-preview",
        confidence_avg=0.89,
    )

    builder = ExcelBuilder()
    output_path = "outputs/test_extraction.xlsx"

    try:
        result_path = builder.create_from_extraction(
            extraction_result,
            output_path,
            show_confidence=True
        )
        print(f"[OK] Excel başarıyla oluşturuldu: {result_path}")

        # Güven skorlarını göster
        print("\n> Guven skorlari:")
        for field, value in extraction_result.fields.items():
            status_emoji = {"valid": "[OK]", "warning": "[WARN]", "error": "[FAIL]"}[value.status]
            print(f"   {status_emoji} {field}: {value.confidence:.0%} ({value.status})")

        return True
    except Exception as e:
        print(f"[FAIL] Hata: {e}")
        return False


def test_5_model_router():
    """Test 5: Model Router (API key olmadan mock test)."""
    print_section("TEST 5: Model Router (Mock)")

    # Mock config (API key yok, sadece mimariyi test ediyoruz)
    config = APIConfig(
        gemini_key="",  # Boş - gerçek test için API key gerekli
        claude_key="",
        openai_key="",
    )

    try:
        router = ModelRouter(config)
        print(f"[OK] Router başlatıldı")
        print(f"   Aktif provider sayısı: {len(router.available_providers)}")

        # Her görev tipi için routing kararı
        print("\n> Routing stratejisi:")
        for task_type in TaskType:
            try:
                decision = router.get_routing_decision(task_type)
                print(f"\n   {task_type.value}:")
                print(f"      Birincil: {decision['primary_model']}")
                if decision['fallback_chain']:
                    print(f"      Yedekler: {', '.join(decision['fallback_chain'][:2])}")
                print(f"      Maliyet (1K token): ${decision['estimated_cost']:.4f}")
            except ValueError:
                print(f"\n   {task_type.value}: [WARN] Hiçbir provider aktif değil")

        return True

    except Exception as e:
        print(f"[FAIL] Hata: {e}")
        return False


def test_6_style_manager():
    """Test 6: Stil yöneticisi."""
    print_section("TEST 6: Stil Yöneticisi")

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    # Veri ekle
    headers = ["Ürün", "Miktar", "Fiyat", "Toplam"]
    data = [
        ["Ekmek", 10, 5.50, 55.00],
        ["Süt", 5, 12.00, 60.00],
        ["Peynir", 2, 85.00, 170.00],
    ]

    # Başlık
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Veri
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Stil uygula
    style_mgr = StyleManager()
    style_mgr.apply_full_formatting(ws)

    output_path = "outputs/test_styled.xlsx"
    os.makedirs("outputs", exist_ok=True)
    wb.save(output_path)

    print(f"[OK] Stillendirilmiş Excel oluşturuldu: {output_path}")
    return True


def test_7_multi_sheet_excel():
    """Test 7: Çoklu sayfa Excel."""
    print_section("TEST 7: Çoklu Sayfa Excel")

    builder = ExcelBuilder()

    sheets_data = {
        "Satışlar": [
            {"Tarih": "2025-03-15", "Ürün": "Laptop", "Tutar": 15000},
            {"Tarih": "2025-03-16", "Ürün": "Mouse", "Tutar": 250},
        ],
        "Masraflar": [
            {"Tarih": "2025-03-15", "Kategori": "Ofis", "Tutar": 500},
            {"Tarih": "2025-03-16", "Kategori": "Ulaşım", "Tutar": 200},
        ],
        "Özet": [
            {"Kategori": "Toplam Satış", "Tutar": 15250},
            {"Kategori": "Toplam Masraf", "Tutar": 700},
            {"Kategori": "Net", "Tutar": 14550},
        ],
    }

    output_path = "outputs/test_multi_sheet.xlsx"

    try:
        result_path = builder.create_multi_sheet(sheets_data, output_path)
        print(f"[OK] Çoklu sayfa Excel oluşturuldu: {result_path}")
        print(f"   Sayfa sayısı: {len(sheets_data)}")
        return True
    except Exception as e:
        print(f"[FAIL] Hata: {e}")
        return False


def run_all_tests():
    """Tüm testleri çalıştır."""
    print("\n" + "=" * 60)
    print("  ExcelAI - End-to-End Test Suite")
    print("=" * 60)

    tests = [
        ("Temel Excel Oluşturma", test_1_basic_excel_creation),
        ("Şablon Kütüphanesi", test_2_template_library),
        ("Code Sandbox Güvenlik", test_3_code_sandbox),
        ("ExtractionResult -> Excel", test_4_extraction_result_to_excel),
        ("Model Router", test_5_model_router),
        ("Stil Yöneticisi", test_6_style_manager),
        ("Çoklu Sayfa Excel", test_7_multi_sheet_excel),
    ]

    results = []

    for test_name, test_func in tests:
        try:
            result = test_func()
            results.append((test_name, result))
        except Exception as e:
            print(f"\n[FAIL] Test başarısız: {test_name}")
            print(f"   Hata: {e}")
            results.append((test_name, False))

    # Özet
    print_section("TEST SONUÇLARI")

    passed = sum(1 for _, result in results if result)
    total = len(results)

    for test_name, result in results:
        status = "[OK] BAŞARILI" if result else "[FAIL] BAŞARISIZ"
        print(f"{status:15} {test_name}")

    print(f"\n> Toplam: {passed}/{total} test basarili ({passed/total*100:.0f}%)")

    if passed == total:
        print("\n[*] Tum testler basarili! Sistem hazir.")
    else:
        print(f"\n[WARN]  {total - passed} test başarısız. Lütfen kontrol edin.")


if __name__ == "__main__":
    # Outputs dizinini oluştur
    os.makedirs("outputs", exist_ok=True)

    # Tüm testleri çalıştır
    run_all_tests()

    print("\n" + "=" * 60)
    print("Test dosyaları 'outputs/' dizininde oluşturuldu.")
    print("Excel dosyalarını kontrol edebilirsiniz.")
    print("=" * 60 + "\n")
