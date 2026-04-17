"""Excel Builder — AI çıktılarını profesyonel Excel dosyalarına dönüştürür."""
import os
from datetime import datetime, date
from decimal import Decimal
from typing import Any, Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from excel_engine.styles import StyleManager
from core.models import ExtractionResult, FieldValue, ColumnType


class ExcelBuilder:
    """
    Excel oluşturma motoru.

    AI'dan gelen yapılandırılmış veriyi alır ve profesyonel
    Excel dosyası oluşturur.
    """

    def __init__(self):
        self.style_manager = StyleManager()

    def create_from_extraction(
        self,
        extraction_result: ExtractionResult,
        output_path: str,
        sheet_name: str = "Veri",
        show_confidence: bool = False,
    ) -> str:
        """
        ExtractionResult'tan Excel oluştur.

        Args:
            extraction_result: AI çıkarma sonucu
            output_path: Çıktı dosya yolu
            sheet_name: Sayfa adı
            show_confidence: Güven skorlarını göster

        Returns:
            Oluşturulan dosya yolu
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel max 31 karakter

        # Başlık satırı
        headers = list(extraction_result.fields.keys())
        if show_confidence:
            headers.append("Güven Skoru")

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Veri satırı
        row_idx = 2
        for col_idx, (field_name, field_value) in enumerate(extraction_result.fields.items(), start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Değer ata
            cell.value = self._convert_value(field_value.value)

            # Güven skoruna göre stil
            if show_confidence:
                style = self.style_manager.create_confidence_style(field_value.confidence)
                cell.font = style["font"]
                cell.fill = style["fill"]
                cell.alignment = style["alignment"]
                cell.border = style["border"]

        # Güven skoru sütunu
        if show_confidence:
            ws.cell(row=row_idx, column=len(headers), value=f"{extraction_result.confidence_avg:.0%}")

        # Formatlama uygula
        self.style_manager.apply_full_formatting(ws)

        # Kaydet
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)

        return output_path

    def create_from_dict_list(
        self,
        data: list[dict[str, Any]],
        output_path: str,
        sheet_name: str = "Veri",
        column_types: Optional[dict[str, ColumnType]] = None,
    ) -> str:
        """
        Dictionary listesinden Excel oluştur.

        Args:
            data: [{"alan1": değer1, "alan2": değer2}, ...]
            output_path: Çıktı dosya yolu
            sheet_name: Sayfa adı
            column_types: Sütun tipleri {"alan": ColumnType}

        Returns:
            Oluşturulan dosya yolu
        """
        if not data:
            raise ValueError("Veri listesi boş")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        # Başlık satırı
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Veri satırları
        for row_idx, record in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = record.get(header)

                # Değer dönüşümü
                cell.value = self._convert_value(value)

                # Tip bazlı stil
                if column_types and header in column_types:
                    is_even = (row_idx % 2) == 0
                    col_type = column_types[header]

                    if col_type == ColumnType.CURRENCY:
                        style = self.style_manager.create_currency_style(is_even)
                    elif col_type == ColumnType.DATE:
                        style = self.style_manager.create_date_style(is_even)
                    elif col_type == ColumnType.NUMBER:
                        style = self.style_manager.create_number_style(is_even)
                    elif col_type == ColumnType.PERCENTAGE:
                        style = self.style_manager.create_percentage_style(is_even)
                    else:
                        style = self.style_manager.create_data_style(is_even)

                    # Stil uygula
                    if "number_format" in style:
                        cell.number_format = style["number_format"]
                    cell.font = style["font"]
                    cell.fill = style["fill"]
                    cell.alignment = style["alignment"]
                    cell.border = style["border"]

        # Formatlama uygula
        self.style_manager.apply_full_formatting(ws)

        # Kaydet
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)

        return output_path

    def create_multi_sheet(
        self,
        sheets: dict[str, list[dict[str, Any]]],
        output_path: str,
    ) -> str:
        """
        Çoklu sayfa içeren Excel oluştur.

        Args:
            sheets: {"Sayfa1": [data], "Sayfa2": [data]}
            output_path: Çıktı dosya yolu

        Returns:
            Oluşturulan dosya yolu
        """
        wb = Workbook()
        wb.remove(wb.active)  # Varsayılan sayfayı kaldır

        for sheet_name, data in sheets.items():
            if not data:
                continue

            ws = wb.create_sheet(title=sheet_name[:31])

            # Başlık
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # Veri
            for row_idx, record in enumerate(data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = self._convert_value(record.get(header))

            # Formatlama
            self.style_manager.apply_full_formatting(ws)

        # Kaydet
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)

        return output_path

    def add_summary_row(
        self,
        ws: Worksheet,
        summary_data: dict[str, Any],
        row_idx: Optional[int] = None,
    ):
        """
        Özet/toplam satırı ekle.

        Args:
            ws: Worksheet
            summary_data: {"Sütun Adı": değer}
            row_idx: Satır numarası (None ise en alta ekle)
        """
        if row_idx is None:
            row_idx = ws.max_row + 1

        # Başlık satırından sütun indekslerini al
        headers = {cell.value: cell.column for cell in ws[1]}

        for col_name, value in summary_data.items():
            if col_name in headers:
                col_idx = headers[col_name]
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = self._convert_value(value)

                # Kalın font + kenarlık
                from openpyxl.styles import Font
                cell.font = Font(bold=True)

    def _convert_value(self, value: Any) -> Any:
        """
        Python değerini Excel uyumlu formata dönüştür.

        Dönüşümler:
        - Decimal → float
        - date/datetime → Python datetime (Excel otomatik formatlar)
        - None → ""
        """
        if value is None:
            return ""
        elif isinstance(value, Decimal):
            return float(value)
        elif isinstance(value, (date, datetime)):
            return value
        elif isinstance(value, bool):
            return value
        elif isinstance(value, (int, float)):
            return value
        else:
            return str(value)


# Yardımcı fonksiyonlar

def quick_excel(
    data: list[dict],
    output_path: str,
    sheet_name: str = "Veri"
) -> str:
    """
    Hızlı Excel oluştur (tek satır kullanım).

    Example:
        quick_excel(
            data=[{"Ad": "Ali", "Soyad": "Yılmaz", "Yaş": 30}],
            output_path="output.xlsx"
        )
    """
    builder = ExcelBuilder()
    return builder.create_from_dict_list(data, output_path, sheet_name)


def extraction_to_excel(
    extraction_result: ExtractionResult,
    output_path: str,
    show_confidence: bool = True
) -> str:
    """
    ExtractionResult'ı direkt Excel'e dönüştür.

    Example:
        extraction_to_excel(result, "fatura.xlsx", show_confidence=True)
    """
    builder = ExcelBuilder()
    return builder.create_from_extraction(
        extraction_result,
        output_path,
        show_confidence=show_confidence
    )
