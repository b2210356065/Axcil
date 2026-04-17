"""Profesyonel Excel stilleri ve formatları."""
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment,
    NamedStyle, numbers
)
from openpyxl.worksheet.worksheet import Worksheet


class StyleManager:
    """
    Excel stil yöneticisi.

    Profesyonel Excel dosyaları için standart stiller:
    - Başlık satırı: kalın, koyu mavi (#1E3A5F), beyaz yazı
    - Zebra şerit (alternatif satır renklendirme)
    - Kenarlıklar
    - Sayı formatları
    """

    # Renk paleti
    COLOR_HEADER_BG = "1E3A5F"      # Koyu mavi
    COLOR_HEADER_TEXT = "FFFFFF"    # Beyaz
    COLOR_ZEBRA_1 = "FFFFFF"        # Beyaz
    COLOR_ZEBRA_2 = "F2F2F2"        # Açık gri
    COLOR_BORDER = "CCCCCC"         # Gri
    COLOR_WARNING = "FFF4E6"        # Açık turuncu
    COLOR_ERROR = "FFE6E6"          # Açık kırmızı
    COLOR_SUCCESS = "E6FFE6"        # Açık yeşil

    @staticmethod
    def create_header_style() -> dict:
        """Başlık satırı stili."""
        return {
            "font": Font(
                name="Calibri",
                size=11,
                bold=True,
                color=StyleManager.COLOR_HEADER_TEXT
            ),
            "fill": PatternFill(
                start_color=StyleManager.COLOR_HEADER_BG,
                end_color=StyleManager.COLOR_HEADER_BG,
                fill_type="solid"
            ),
            "alignment": Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            ),
            "border": StyleManager._create_border()
        }

    @staticmethod
    def create_data_style(is_even_row: bool = False) -> dict:
        """Veri satırı stili (zebra şerit)."""
        bg_color = StyleManager.COLOR_ZEBRA_2 if is_even_row else StyleManager.COLOR_ZEBRA_1

        return {
            "font": Font(name="Calibri", size=10),
            "fill": PatternFill(
                start_color=bg_color,
                end_color=bg_color,
                fill_type="solid"
            ),
            "alignment": Alignment(
                horizontal="left",
                vertical="center"
            ),
            "border": StyleManager._create_border(style="thin")
        }

    @staticmethod
    def create_number_style(is_even_row: bool = False) -> dict:
        """Sayısal veri stili (sağa hizalı, binlik ayırıcılı)."""
        style = StyleManager.create_data_style(is_even_row)
        style["alignment"] = Alignment(horizontal="right", vertical="center")
        style["number_format"] = "#,##0"
        return style

    @staticmethod
    def create_currency_style(is_even_row: bool = False, currency_symbol: str = "₺") -> dict:
        """Para birimi stili."""
        style = StyleManager.create_data_style(is_even_row)
        style["alignment"] = Alignment(horizontal="right", vertical="center")
        style["number_format"] = f'#,##0.00 "{currency_symbol}"'
        return style

    @staticmethod
    def create_percentage_style(is_even_row: bool = False) -> dict:
        """Yüzde stili."""
        style = StyleManager.create_data_style(is_even_row)
        style["alignment"] = Alignment(horizontal="right", vertical="center")
        style["number_format"] = "0%"
        return style

    @staticmethod
    def create_date_style(is_even_row: bool = False) -> dict:
        """Tarih stili (DD.MM.YYYY)."""
        style = StyleManager.create_data_style(is_even_row)
        style["alignment"] = Alignment(horizontal="center", vertical="center")
        style["number_format"] = "DD.MM.YYYY"
        return style

    @staticmethod
    def create_confidence_style(confidence: float) -> dict:
        """
        Güven skoruna göre arka plan rengi.

        >= 0.90: Yeşil (otomatik kabul)
        0.70-0.90: Sarı (kullanıcı onayı)
        < 0.70: Kırmızı (manuel düzeltme)
        """
        if confidence >= 0.90:
            bg_color = StyleManager.COLOR_SUCCESS
        elif confidence >= 0.70:
            bg_color = StyleManager.COLOR_WARNING
        else:
            bg_color = StyleManager.COLOR_ERROR

        return {
            "font": Font(name="Calibri", size=10),
            "fill": PatternFill(
                start_color=bg_color,
                end_color=bg_color,
                fill_type="solid"
            ),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": StyleManager._create_border(style="thin")
        }

    @staticmethod
    def _create_border(style: str = "medium") -> Border:
        """Kenarlık oluştur."""
        side = Side(
            style=style,
            color=StyleManager.COLOR_BORDER
        )
        return Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    def apply_header_row(ws: Worksheet, row: int = 1):
        """Başlık satırına stil uygula."""
        header_style = StyleManager.create_header_style()

        for cell in ws[row]:
            cell.font = header_style["font"]
            cell.fill = header_style["fill"]
            cell.alignment = header_style["alignment"]
            cell.border = header_style["border"]

    @staticmethod
    def apply_zebra_striping(ws: Worksheet, start_row: int = 2, end_row: int = None):
        """Alternatif satır renklendirme (zebra şerit)."""
        if end_row is None:
            end_row = ws.max_row

        for row_idx in range(start_row, end_row + 1):
            is_even = (row_idx % 2) == 0
            data_style = StyleManager.create_data_style(is_even)

            for cell in ws[row_idx]:
                if cell.value is not None:  # Sadece dolu hücrelere
                    cell.font = data_style["font"]
                    cell.fill = data_style["fill"]
                    cell.alignment = data_style["alignment"]
                    cell.border = data_style["border"]

    @staticmethod
    def auto_adjust_column_width(ws: Worksheet, min_width: int = 10, max_width: int = 50):
        """Sütun genişliklerini içeriğe göre otomatik ayarla."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def freeze_header(ws: Worksheet):
        """Başlık satırını dondur (scroll yaparken sabit kalır)."""
        ws.freeze_panes = "A2"

    @staticmethod
    def set_print_options(ws: Worksheet):
        """Yazdırma ayarları."""
        ws.page_setup.orientation = "landscape"  # Yatay
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # Yüksekliği sınırlama
        ws.print_options.horizontalCentered = True

    @staticmethod
    def apply_full_formatting(
        ws: Worksheet,
        has_header: bool = True,
        freeze_header: bool = True,
        zebra_striping: bool = True,
        auto_width: bool = True,
        print_setup: bool = True
    ):
        """
        Tüm formatlamayı tek seferde uygula.

        Args:
            ws: Worksheet
            has_header: Başlık satırı var mı?
            freeze_header: Başlık dondurusu
            zebra_striping: Alternatif satır renklendirme
            auto_width: Otomatik sütun genişliği
            print_setup: Yazdırma ayarları
        """
        if has_header:
            StyleManager.apply_header_row(ws, row=1)
            if freeze_header:
                StyleManager.freeze_header(ws)

        if zebra_striping:
            start_row = 2 if has_header else 1
            StyleManager.apply_zebra_striping(ws, start_row=start_row)

        if auto_width:
            StyleManager.auto_adjust_column_width(ws)

        if print_setup:
            StyleManager.set_print_options(ws)


# Önceden tanımlanmış stiller (openpyxl NamedStyle olarak)
def register_named_styles(workbook):
    """Workbook'a önceden tanımlanmış stilleri kaydet."""
    styles = {
        "header": NamedStyle(name="header"),
        "currency_tl": NamedStyle(name="currency_tl"),
        "date_tr": NamedStyle(name="date_tr"),
    }

    # Header style
    styles["header"].font = Font(bold=True, color="FFFFFF")
    styles["header"].fill = PatternFill(start_color="1E3A5F", fill_type="solid")
    styles["header"].alignment = Alignment(horizontal="center", vertical="center")

    # Currency TL style
    styles["currency_tl"].number_format = '#,##0.00 "₺"'
    styles["currency_tl"].alignment = Alignment(horizontal="right")

    # Date TR style
    styles["date_tr"].number_format = "DD.MM.YYYY"
    styles["date_tr"].alignment = Alignment(horizontal="center")

    # Workbook'a ekle
    for style in styles.values():
        if style.name not in workbook.named_styles:
            workbook.add_named_style(style)
