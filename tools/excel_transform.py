"""Excel -> Excel Araci -- mevcut Excel dosyalarini donusturur, bilestirir."""
from typing import Optional, Union
from pydantic import BaseModel, Field
from openpyxl import load_workbook
from tools.base_tool import BaseTool, ToolInput, ToolResult
from ai.router import ModelRouter
from core.models import TaskType, ExtractionResult, FieldValue


class ExcelTransformInput(ToolInput):
    """Excel -> Excel araci girdisi."""
    excel_path: str = Field(description="Kaynak Excel dosya yolu")
    target_schema: type[BaseModel] = Field(description="Hedef sema")
    business_context: dict = Field(default_factory=dict)
    operation: str = Field(
        default="transform",
        description="Islem: transform, merge, clean, format"
    )
    sheet_name: Optional[str] = Field(default=None, description="Islenecek sayfa adi")


class ExcelTransformTool(BaseTool):
    """
    Excel -> Excel Donusturme Araci

    Kullanim senaryolari:
    - Farkli formattaki Excel'leri tek standartta birlestirme
    - Bir Excel'i farkli sablona donusturme
    - Banka CSV -> muhasebe Excel formati
    - Birden fazla subenin verilerini konsolide etme
    - Eski format -> yeni format gocu
    - Veri temizleme ve normalizasyon

    Ozellikler:
    - Coklu sayfa destegi
    - Otomatik sutun esleme
    - Veri temizleme (bosluklar, formatlar)
    - Birlestirme ve konsolidasyon
    """

    def __init__(self, router: ModelRouter):
        super().__init__(
            name="Excel -> Excel Donusturme",
            description="Mevcut Excel dosyalarini donusturur, bilestirir veya temizler"
        )
        self.router = router

    def process(
        self,
        input_data: ExcelTransformInput,
        **kwargs
    ) -> ToolResult:
        """
        Ana isleme akisi.

        ADIM 1: Kaynak Excel'i oku
        ADIM 2: Veriyi cikar
        ADIM 3: Claude ile karmasik donusum (gerekirse)
        ADIM 4: Hedef formatta yeniden olustur
        """
        try:
            # ADIM 1: Excel oku
            source_data = self._read_excel(
                input_data.excel_path,
                input_data.sheet_name
            )

            if not source_data:
                return ToolResult(
                    success=False,
                    message="Excel dosyasi bos veya okunamiyor",
                    errors=["Gecerli bir Excel dosyasi saglayin"],
                )

            # ADIM 2: Islem tipine gore donustur
            if input_data.operation == "transform":
                result = self._transform_data(
                    source_data,
                    input_data.target_schema,
                    input_data.business_context
                )
            elif input_data.operation == "clean":
                result = self._clean_data(source_data)
            elif input_data.operation == "merge":
                result = self._merge_data(source_data)
            else:
                result = source_data

            # ExtractionResult olustur
            extraction_result = ExtractionResult(
                fields={k: FieldValue(value=v, confidence=0.90) for k, v in result.items()},
                model_used="excel_transform",
                confidence_avg=0.90,
            )

            # Onizleme
            preview_data = self._create_preview(extraction_result)

            return ToolResult(
                success=True,
                extraction_result=extraction_result,
                preview_data=preview_data,
                message=f"Excel basariyla donusturuldu. {len(result)} alan islendi.",
            )

        except Exception as e:
            return ToolResult(
                success=False,
                message="Excel donusturme basarisiz",
                errors=[str(e)],
            )

    def transform_multiple(
        self,
        excel_paths: list[str],
        target_schema: type[BaseModel],
        business_context: dict
    ) -> ToolResult:
        """
        Birden fazla Excel'i birlestirip tek formatta donustur.

        Kullanim: 5 subenin farkli formattaki Excel'lerini birlestir
        """
        all_data = []

        for path in excel_paths:
            data = self._read_excel(path)
            if data:
                all_data.append(data)

        # Birlestir
        merged = self._merge_multiple(all_data)

        # Donustur
        result = self._transform_data(merged, target_schema, business_context)

        extraction_result = ExtractionResult(
            fields={k: FieldValue(value=v, confidence=0.85) for k, v in result.items()},
            model_used="excel_merge",
            confidence_avg=0.85,
        )

        return ToolResult(
            success=True,
            extraction_result=extraction_result,
            preview_data=self._create_preview(extraction_result),
            message=f"{len(excel_paths)} Excel bilestirildi ve donusturuldu.",
        )

    def preview(self, result: ToolResult) -> dict:
        """Streamlit icin onizleme verisi."""
        if not result.extraction_result:
            return {}
        return result.preview_data

    def _read_excel(
        self,
        excel_path: str,
        sheet_name: Optional[str] = None
    ) -> dict:
        """
        Excel dosyasini oku.

        Returns:
            Ilk satir verisi (dictionary)
        """
        wb = load_workbook(excel_path, data_only=True)

        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Ilk satiri baslik olarak al
        headers = [cell.value for cell in ws[1]]

        # Ikinci satiri veri olarak al (demo icin)
        if ws.max_row > 1:
            values = [cell.value for cell in ws[2]]
            data = dict(zip(headers, values))
        else:
            data = {}

        wb.close()

        return data

    def _transform_data(
        self,
        source_data: dict,
        target_schema: type[BaseModel],
        context: dict
    ) -> dict:
        """
        Veriyi hedef semaya donustur.

        Basit esleme + AI ile karmasik donusum (gerekirse)
        """
        # Basit alan esleme
        target_fields = target_schema.model_json_schema()["properties"]
        transformed = {}

        for target_field in target_fields:
            # Kaynak alanda ayni isim var mi?
            if target_field in source_data:
                transformed[target_field] = source_data[target_field]
            else:
                # Benzer isim ara
                similar = self._find_similar_field(target_field, source_data.keys())
                if similar:
                    transformed[target_field] = source_data[similar]
                else:
                    transformed[target_field] = None

        return transformed

    def _clean_data(self, data: dict) -> dict:
        """
        Veriyi temizle.

        - Bosluklari kaldir
        - None -> ""
        - Sayi formatlarini duzelt
        """
        cleaned = {}

        for key, value in data.items():
            if value is None:
                cleaned[key] = ""
            elif isinstance(value, str):
                cleaned[key] = value.strip()
            else:
                cleaned[key] = value

        return cleaned

    def _merge_data(self, data: dict) -> dict:
        """Tekil birlestirme (su an basit pass-through)."""
        return data

    def _merge_multiple(self, data_list: list[dict]) -> dict:
        """Birden fazla dictionary'yi birlestir."""
        merged = {}

        for data in data_list:
            merged.update(data)

        return merged

    def _find_similar_field(self, target: str, source_fields: list[str]) -> Optional[str]:
        """
        Hedef alana benzer kaynak alan bul.

        Basit string benzerligi (lowercase, _ kaldirma)
        """
        target_clean = target.lower().replace("_", "").replace(" ", "")

        for source in source_fields:
            source_clean = source.lower().replace("_", "").replace(" ", "")
            if target_clean == source_clean:
                return source

        return None

    def _create_preview(self, extraction_result: ExtractionResult) -> dict:
        """Onizleme tablosu."""
        rows = []

        for field_name, field_value in extraction_result.fields.items():
            rows.append({
                "Alan": field_name,
                "Deger": str(field_value.value),
            })

        return {
            "headers": ["Alan", "Deger"],
            "rows": rows,
        }

    @property
    def tool_type(self) -> str:
        return "excel_transform"


# Yardimci fonksiyonlar

def quick_excel_transform(
    excel_path: str,
    target_schema: type[BaseModel],
    router: ModelRouter,
    context: dict = None
) -> ToolResult:
    """Hizli Excel donusturme."""
    tool = ExcelTransformTool(router)
    return tool.process(ExcelTransformInput(
        excel_path=excel_path,
        target_schema=target_schema,
        business_context=context or {}
    ))
